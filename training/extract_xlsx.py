"""
Extract historical training data from local xlsx archives.

Three xlsx files cover the historical record:
  - 2023 Scripts.xlsx
  - 2024 Scripts.xlsx
  - 2025 Scripts.xlsx

Two schemas live across these files:

  OLD (2023, 2024):
    A=Date, B=Type, C=Studio, D=Talent(s), E=Dynamic/Setting, F=Wardrobe,
    G=Plot/Background, H=Shoot Concept, I=Suggested Title, J=Props

  NEW (2025, matches the live 2026 Scripts Sheet):
    A=Date, B=Studio, C=Location, D=Scene, E=Female Model, F=Male Model,
    G=Theme, H=Wardrobe Female, I=Wardrobe Male, J=Plot, K=Title, L=Props

The extractor auto-detects the schema by header row, normalizes studios
(FPVR/VRH/VRA/NJOI → canonical), normalizes scene types
(BG-CP/Solo/etc. → BGCP/SOLO), and splits combined talent columns into
female/male leads.

Output: same JSONL schema as extract_scripts.py, so the two datasets can
be concatenated for training.

Usage:
    python3 training/extract_xlsx.py [paths...] --out training/data/historical_dataset.jsonl

If no paths are given, defaults to the three files in ~/Downloads/.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Schema detection
# ---------------------------------------------------------------------------

OLD_HEADER_KEYS = ("Type", "Talent")          # Old schema has a "Type" column and "Talent(s)"
NEW_HEADER_KEYS = ("Female Model", "Male Model")  # New schema has split female/male columns


def detect_schema(header: tuple) -> str:
    cells = [str(c).strip() if c else "" for c in header]
    has_type = any(k in cells for k in ["Type"])
    has_female = any("Female Model" in c for c in cells)
    if has_female:
        return "new"
    if has_type:
        return "old"
    return "unknown"


# ---------------------------------------------------------------------------
# Studio normalization
# ---------------------------------------------------------------------------

# Older sheets used "FPVR - City, ST" — split off the city into destination
FPVR_LOCATION_RE = re.compile(r"^FPVR\s*[-—]\s*(.+)$", re.IGNORECASE)

# Typo studios that have stray digits or extra suffixes attached
# "FPVR93", "FPVR94", "VRH458", etc. Strip the trailing junk.
STUDIO_TYPO_RE = re.compile(r"^(FPVR|VRH|VRA|NJOI|NNJOI)[\s\d]+$", re.IGNORECASE)

# "FPVR / XPVR", "XPVR / VRH" — composite studio cells (multi-studio shoots).
# Take the first canonical studio that appears.
STUDIO_COMPOSITE_RE = re.compile(r"\bFPVR\b|\bVRH\b|\bVRA\b|\bNJOI\b|\bNNJOI\b", re.IGNORECASE)

STUDIO_ALIASES = {
    "fpvr":       "FuckPassVR",
    "fuckpassvr": "FuckPassVR",
    "vrh":        "VRHush",
    "vrhush":     "VRHush",
    "vra":        "VRAllure",
    "vrallure":   "VRAllure",
    "njoi":       "NaughtyJOI",
    "nnjoi":      "NaughtyJOI",
    "naughtyjoi": "NaughtyJOI",
}

# Anything else (BJN, XPVR, BlowJobNow, etc.) — different product line, skip
SKIP_STUDIOS = {"bjn", "xpvr", "blowjobnow", "bjvr"}


def normalize_studio_old(raw: str) -> tuple[str, str]:
    """
    Old-schema studio cell. Returns (canonical_studio, destination_or_empty).

    "FPVR - New Orleans, LA" → ("FuckPassVR", "New Orleans, LA")
    "VRH" → ("VRHush", "")
    "FPVR94" → ("FuckPassVR", "")  (strips typo trailing digits)
    "FPVR / XPVR" → ("FuckPassVR", "")  (multi-studio — pick first canonical)
    "BJN" → ("", "") — caller skips
    """
    raw = (raw or "").strip()
    if not raw:
        return "", ""

    # FPVR with destination
    m = FPVR_LOCATION_RE.match(raw)
    if m:
        return "FuckPassVR", m.group(1).strip()

    # Typo with trailing digits / spaces (FPVR93, VRH458)
    m = STUDIO_TYPO_RE.match(raw.replace(" ", ""))
    if m:
        token = m.group(1).lower()
        if token in STUDIO_ALIASES:
            return STUDIO_ALIASES[token], ""

    key = raw.lower().strip()
    if key in SKIP_STUDIOS:
        return "", ""
    if key in STUDIO_ALIASES:
        return STUDIO_ALIASES[key], ""

    # Composite studio cells like "FPVR / XPVR" — pick first canonical token
    composite_match = STUDIO_COMPOSITE_RE.search(raw)
    if composite_match:
        token = composite_match.group(0).lower()
        if token in STUDIO_ALIASES:
            return STUDIO_ALIASES[token], ""

    # Unknown — skip
    return "", ""


def normalize_studio_new(raw: str) -> str:
    raw = (raw or "").strip()
    if not raw:
        return ""
    return STUDIO_ALIASES.get(raw.lower(), raw)


# ---------------------------------------------------------------------------
# Scene type normalization
# ---------------------------------------------------------------------------

SCENE_TYPE_ALIASES = {
    "bg":              "BG",
    "bgcp":            "BGCP",
    "bg-cp":           "BGCP",
    "bg - cp":         "BGCP",
    "bg cp":           "BGCP",
    "bg-cp (faked)":   "BGCP",
    "bg - cp (faked)": "BGCP",
    "bgcp (faked)":    "BGCP",
    "bg-fake cp":      "BGCP",
    "bga-cp":          "BGCP",   # BG + anal + creampie — collapse to BGCP for training
    "bga":             "BG",
    "bg-a":            "BG",
    "cp":              "BGCP",
    "solo":            "SOLO",
    "joi":             "JOI",
    "bj":              "BJ",     # blowjob-only — BJN studio, will be skipped upstream
}


def normalize_scene_type(raw: str, studio: str) -> str:
    if studio == "VRAllure":
        return "SOLO"
    if studio == "NaughtyJOI":
        return "JOI"
    key = (raw or "").strip().lower()
    if key in SCENE_TYPE_ALIASES:
        return SCENE_TYPE_ALIASES[key]
    # Fallback — strip parens/anal markers and try again
    cleaned = re.sub(r"\s*\(.*?\)\s*", "", key).strip()
    if cleaned in SCENE_TYPE_ALIASES:
        return SCENE_TYPE_ALIASES[cleaned]
    return (raw or "BG").strip().upper()


# ---------------------------------------------------------------------------
# Talent parsing (old schema)
# ---------------------------------------------------------------------------

# Old "Talent(s)" column has formats:
#   "Haley Spades, Austin Pierce"        → female + male
#   "Haley Spades / Mike Mancini"        → female + male
#   "Haley Spades"                       → female only (solo)
#   "Mike Mancini / Sky Pierce"          → ??? (rare — male first)
#
# For this dataset, female is named first by convention. When in doubt the
# first name in the cell is the female lead.

TALENT_SPLIT_RE = re.compile(r"\s*[/,&]\s*|\s+and\s+", re.IGNORECASE)


def parse_talent(raw: str) -> tuple[str, str]:
    """Return (female, male). Male defaults to 'POV' for solo entries."""
    raw = (raw or "").strip()
    if not raw:
        return "", ""
    parts = [p.strip() for p in TALENT_SPLIT_RE.split(raw) if p.strip()]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], "POV"
    return parts[0], parts[1]


# ---------------------------------------------------------------------------
# Output formatters (matches extract_scripts.py — keep these compatible!)
# ---------------------------------------------------------------------------

def format_input(studio: str, scene_type: str, female: str, male: str) -> str:
    parts = [
        f"Studio: {studio}",
        f"Scene Type: {scene_type or 'BG'}",
        f"Female Talent: {female}",
    ]
    if studio != "VRAllure":
        parts.append(f"Male Talent: {male or 'POV'}")
    return "\n".join(parts)


def format_output(
    *, studio: str, theme: str, plot: str, location: str,
    props: str, wardrobe_f: str, wardrobe_m: str,
) -> str:
    sections = [
        ("THEME", theme),
        ("PLOT", plot),
        ("SHOOT LOCATION", location),
        ("PROPS", props),
        ("WARDROBE - FEMALE", wardrobe_f),
    ]
    if studio != "VRAllure":
        sections.append(("WARDROBE - MALE", wardrobe_m))
    return "\n\n".join(f"{label}: {value}" for label, value in sections if value)


# ---------------------------------------------------------------------------
# Per-row extraction
# ---------------------------------------------------------------------------

def extract_old_row(row: tuple, source_tab: str, source_row: int, source_file: str) -> dict | None:
    """Old schema: Date, Type, Studio, Talent(s), Dynamic/Setting, Wardrobe, Plot/Background, Shoot Concept, Suggested Title, Props"""
    padded = list(row) + [None] * (10 - len(row))
    raw_date, raw_type, raw_studio, raw_talent, raw_setting, raw_wardrobe, raw_plot, raw_concept, raw_title, raw_props = padded[:10]

    studio, destination = normalize_studio_old(str(raw_studio or ""))
    if not studio:
        return None  # SKIP studios (BJN, XPVR, etc.) or empty

    female, male = parse_talent(str(raw_talent or ""))
    if not female:
        return None  # need at least a female lead

    plot = str(raw_plot or "").strip()
    if not plot:
        return None  # need a plot to be training-worthy

    scene_type = normalize_scene_type(str(raw_type or ""), studio)
    title = str(raw_title or "").strip()
    # Old schema has no separate Theme — fall back to Suggested Title or Shoot Concept
    theme = title or str(raw_concept or "").strip()
    setting = str(raw_setting or "").strip()
    # Combined Wardrobe → put in wardrobe_f (older sheets are typically female-focused)
    wardrobe = str(raw_wardrobe or "").strip()
    props = str(raw_props or "").strip()

    shoot_date = ""
    if isinstance(raw_date, datetime):
        shoot_date = raw_date.strftime("%Y-%m-%d")
    elif raw_date:
        shoot_date = str(raw_date).strip()

    return {
        "source_file":  source_file,
        "source_tab":   source_tab,
        "source_row":   source_row,
        "schema":       "old",
        "studio":       studio,
        "scene_type":   scene_type,
        "female":       female,
        "male":         male,
        "destination":  destination,
        "shoot_date":   shoot_date,
        "location":     setting,
        "title":        title,
        "theme":        theme,
        "plot":         plot,
        "wardrobe_f":   wardrobe,
        "wardrobe_m":   "",
        "props":        props,
        "status":       "",
        "input":  format_input(studio, scene_type, female, male),
        "output": format_output(
            studio=studio,
            theme=theme, plot=plot, location=setting,
            props=props, wardrobe_f=wardrobe, wardrobe_m="",
        ),
    }


def extract_new_row(row: tuple, source_tab: str, source_row: int, source_file: str) -> dict | None:
    """New schema (matches 2026 live sheet)."""
    padded = list(row) + [None] * (12 - len(row))
    raw_date, raw_studio, raw_location, raw_scene, raw_female, raw_male, raw_theme, raw_wf, raw_wm, raw_plot, raw_title, raw_props = padded[:12]

    studio = normalize_studio_new(str(raw_studio or ""))
    if not studio or studio.lower() in SKIP_STUDIOS:
        return None
    female = str(raw_female or "").strip()
    if not female:
        return None
    plot = str(raw_plot or "").strip()
    theme = str(raw_theme or "").strip()
    if not plot or not theme:
        return None

    male = str(raw_male or "POV").strip() or "POV"
    scene_type = normalize_scene_type(str(raw_scene or ""), studio)
    location = str(raw_location or "").strip()
    title = str(raw_title or "").strip()
    wardrobe_f = str(raw_wf or "").strip()
    wardrobe_m = str(raw_wm or "").strip()
    props = str(raw_props or "").strip()

    shoot_date = ""
    if isinstance(raw_date, datetime):
        shoot_date = raw_date.strftime("%Y-%m-%d")
    elif raw_date:
        shoot_date = str(raw_date).strip()

    return {
        "source_file":  source_file,
        "source_tab":   source_tab,
        "source_row":   source_row,
        "schema":       "new",
        "studio":       studio,
        "scene_type":   scene_type,
        "female":       female,
        "male":         male,
        "destination":  "",
        "shoot_date":   shoot_date,
        "location":     location,
        "title":        title,
        "theme":        theme,
        "plot":         plot,
        "wardrobe_f":   wardrobe_f,
        "wardrobe_m":   wardrobe_m,
        "props":        props,
        "status":       "",
        "input":  format_input(studio, scene_type, female, male),
        "output": format_output(
            studio=studio,
            theme=theme, plot=plot, location=location,
            props=props, wardrobe_f=wardrobe_f, wardrobe_m=wardrobe_m,
        ),
    }


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

SKIP_TAB_NAMES = {"Cancellations", "Template", "Sheet1"}


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument(
        "files",
        nargs="*",
        type=Path,
        default=[
            Path.home() / "Downloads" / "2023 Scripts.xlsx",
            Path.home() / "Downloads" / "2024 Scripts.xlsx",
            Path.home() / "Downloads" / "2025 Scripts.xlsx",
        ],
    )
    p.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).parent / "data" / "historical_dataset.jsonl",
    )
    args = p.parse_args()

    args.out.parent.mkdir(parents=True, exist_ok=True)

    per_file = Counter()
    per_studio = Counter()
    per_scene = Counter()
    skipped = Counter()
    total = 0

    with args.out.open("w", encoding="utf-8") as fh:
        for xlsx_path in args.files:
            if not xlsx_path.exists():
                print(f"  ! {xlsx_path}: not found, skipping", file=sys.stderr)
                continue
            print(f"\n=== {xlsx_path.name} ===")
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                if sheet_name in SKIP_TAB_NAMES:
                    continue
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                schema = detect_schema(rows[0])
                if schema == "unknown":
                    print(f"  ? {sheet_name}: schema not recognized, skipping")
                    continue

                tab_count = 0
                for row_idx, row in enumerate(rows[1:], start=2):
                    if not row or not any(str(c).strip() if c else "" for c in row):
                        continue
                    if schema == "old":
                        record = extract_old_row(row, sheet_name, row_idx, xlsx_path.name)
                    else:
                        record = extract_new_row(row, sheet_name, row_idx, xlsx_path.name)
                    if record is None:
                        skipped["filtered"] += 1
                        continue
                    fh.write(json.dumps(record, ensure_ascii=False) + "\n")
                    total += 1
                    tab_count += 1
                    per_file[xlsx_path.name] += 1
                    per_studio[record["studio"]] += 1
                    per_scene[record["scene_type"]] += 1

                print(f"  • {sheet_name:18s} [{schema}]   {tab_count} rows kept")

            wb.close()

    print(f"\nWrote {total} records to {args.out}")
    print("\n--- Per file ---")
    for f, c in per_file.most_common():
        print(f"  {f:30s}  {c}")
    print("\n--- Per studio ---")
    for s, c in per_studio.most_common():
        print(f"  {s:15s}  {c}")
    print("\n--- Per scene type ---")
    for s, c in per_scene.most_common():
        print(f"  {s:8s}  {c}")
    if skipped:
        print(f"\nSkipped (filtered): {skipped['filtered']}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
