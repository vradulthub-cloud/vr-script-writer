#!/usr/bin/env python3
"""
calculate_avg_rates.py
======================
Reads all Shoot Budget Excel files (2021–present) and for every model:
  • Computes AVG Rate  → writes to "AVG Rate" column
  • Builds booking history → writes to "Dates Booked" column

Dates Booked format:  "8x · Jan 2025"
  (total unique shoot days  ·  most recent booking month)

Budget files live in BUDGET_DIR — drop new yearly files there as they arrive.

Usage:
    python3 /Users/andrewninn/Scripts/calculate_avg_rates.py

Runs automatically on the 1st of every month via the scheduled task.
"""

import re
import gspread
import openpyxl
from pathlib import Path
from collections import defaultdict
from datetime import datetime, date
from google.oauth2.service_account import Credentials

# ── Config ────────────────────────────────────────────────────────────────────

SPREADSHEET_ID       = "1Dxrh0UZNqoBt6otZqsU85fxz9z-dt0csSCV9sGdobKw"
SERVICE_ACCOUNT_FILE = Path(__file__).parent / "service_account.json"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

BUDGET_DIR  = Path(__file__).parent / "shoot_budgets"
HEADER_ROW  = 3   # 1-indexed row with column names in the booking sheet
DATA_START  = 4   # 1-indexed first data row

SKIP_SHEETS = {"cancellations", "template", "flights notes", "summary", "notes"}

# Column name variants across 2021–2026
DATE_COLS   = {"date"}
F_NAME_COLS = {"f talent", "f1 talent", "f2 talent", "f3 talent"}
F_RATE_COLS = {"f1 amount", "f2 amount", "f3 amount", "f amount",
               "f1 rate",  "f rate"}
M_NAME_COLS = {"m talent"}
M_RATE_COLS = {"m amount", "m rate"}


# ── Auth ──────────────────────────────────────────────────────────────────────

def get_gspread():
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return gspread.authorize(creds)


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize(name: str) -> str:
    """Normalize a talent name for matching.
    Strips parenthetical notes like '(A - CP)', '(Solo)', etc. so that
    'Lauren Phillips (A - CP)' matches 'Lauren Phillips'.
    """
    s = str(name).strip()
    s = re.sub(r"\s*\(.*?\)", "", s)   # remove parenthetical suffixes
    return re.sub(r"\s+", " ", s.strip().lower())


def parse_rate(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(str(val).replace(",", "").replace("$", "").strip())
        return f if f > 50 else None
    except (ValueError, TypeError):
        return None


def parse_date(val) -> date | None:
    """Convert openpyxl cell value to a Python date, or None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def fmt_date(d: date) -> str:
    """'Jan 2025'"""
    return d.strftime("%b %Y")


def fmt_dates_booked(dates: list[date]) -> str:
    """'8x · Jan 2025'  — unique shoot days, most recent month"""
    unique = sorted(set(dates))
    if not unique:
        return ""
    count = len(unique)
    last  = unique[-1]
    return f"{count}x · {fmt_date(last)}"


# ── Budget parsing ────────────────────────────────────────────────────────────

def parse_sheet(ws) -> list[tuple[str, float | None, date | None]]:
    """
    Return [(name, rate_or_None, date_or_None), ...] for every talent row.
    Rows where rate is 0 / blank still get captured for date tracking
    only if they have a real date and a real rate on a companion row.
    We capture (name, rate, date) and let the caller decide.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Locate header row
    header_idx = None
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c and str(c).strip()]
        if len(non_empty) >= 3:
            header_idx = i
            break
    if header_idx is None:
        return []

    headers = [normalize(c) if c else "" for c in rows[header_idx]]

    # Date is always column 0, but the header label is often a junk value (number, letter, blank).
    # Try to find it by header name first; fall back to column 0.
    date_idx  = next((i for i, h in enumerate(headers) if h in DATE_COLS), 0)
    f_names   = [i for i, h in enumerate(headers) if h in F_NAME_COLS]
    f_rates   = [i for i, h in enumerate(headers) if h in F_RATE_COLS]
    m_names   = [i for i, h in enumerate(headers) if h in M_NAME_COLS]
    m_rates   = [i for i, h in enumerate(headers) if h in M_RATE_COLS]

    if not (f_names or m_names):
        return []

    results = []

    def get(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        first = str(row[0]).strip().lower() if row[0] else ""
        if first in ("break", "b", ""):
            continue

        row_date = parse_date(get(row, date_idx))

        # Female talent pairs — pair by position, ensuring every name gets a rate slot.
        # Pad rates with None if there are more names than rate cols.
        padded_f_rates = f_rates + [None] * max(0, len(f_names) - len(f_rates))
        for ni, ri in zip(f_names, padded_f_rates):
            name = str(get(row, ni) or "").strip()
            rate = parse_rate(get(row, ri))
            if name and len(name) > 2 and rate:
                results.append((name, rate, row_date))

        # Male talent
        padded_m_rates = m_rates + [None] * max(0, len(m_names) - len(m_rates))
        for ni, ri in zip(m_names, padded_m_rates):
            name = str(get(row, ni) or "").strip()
            rate = parse_rate(get(row, ri))
            if name and len(name) > 2 and rate:
                results.append((name, rate, row_date))

    return results


def build_histories():
    """
    Scan all budget files.
    Returns:
      rate_history  : {normalized_name: [rate, ...]}
      date_history  : {normalized_name: [date, ...]}   (unique dates)
      canonical     : {normalized_name: display_name}
    """
    rate_history  = defaultdict(list)
    date_history  = defaultdict(list)
    canonical     = {}

    files = sorted(BUDGET_DIR.glob("*.xlsx"))
    if not files:
        print(f"  WARNING: No .xlsx files found in {BUDGET_DIR}")
        return rate_history, date_history, canonical

    for fpath in files:
        print(f"  Reading {fpath.name}...")
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception as e:
            print(f"    ERROR opening {fpath.name}: {e}")
            continue

        for sname in wb.sheetnames:
            if normalize(sname) in SKIP_SHEETS:
                continue
            triples = parse_sheet(wb[sname])
            for name, rate, dt in triples:
                key = normalize(name)
                rate_history[key].append(rate)
                if dt:
                    date_history[key].append(dt)
                if key not in canonical:
                    canonical[key] = name

        wb.close()

    return rate_history, date_history, canonical


# ── Rate computation ──────────────────────────────────────────────────────────

def compute_avg_rate(rates: list[float]) -> int:
    """Integer average rounded to nearest $50."""
    if not rates:
        return 0
    return int(round(sum(rates) / len(rates) / 50) * 50)


def rate_summary(rates, dates) -> str:
    if not rates:
        return "no data"
    avg = sum(rates) / len(rates)
    mn, mx = min(rates), max(rates)
    bookings = len(set(dates)) if dates else len(rates)
    return f"${avg:,.0f} avg  (${mn:,.0f}–${mx:,.0f}, {bookings} bookings)"


# ── Sheet update ──────────────────────────────────────────────────────────────

def update_sheet(gc, rate_hist, date_hist, canonical) -> int:
    ss      = gc.open_by_key(SPREADSHEET_ID)
    updated = 0
    no_data = []

    for ws in ss.worksheets():
        tab = ws.title
        try:
            headers = ws.row_values(HEADER_ROW)
        except Exception:
            continue

        col_map = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}
        if "Name" not in col_map:
            continue

        name_col        = col_map["Name"]
        avg_rate_col    = col_map.get("AVG Rate")
        dates_booked_col = col_map.get("Dates Booked")

        if not avg_rate_col and not dates_booked_col:
            print(f"  [{tab}] No AVG Rate or Dates Booked columns — skipping")
            continue

        all_names = ws.col_values(name_col)
        updates   = []

        for row_idx, name in enumerate(all_names, start=1):
            if row_idx < DATA_START:
                continue
            name = str(name).strip()
            if not name:
                continue

            key   = normalize(name)
            rates = rate_hist.get(key, [])
            dates = date_hist.get(key, [])

            if not rates and not dates:
                no_data.append(f"{tab}: {name}")
                continue

            if avg_rate_col and rates:
                avg = compute_avg_rate(rates)
                updates.append({
                    "range":  gspread.utils.rowcol_to_a1(row_idx, avg_rate_col),
                    "values": [[avg]]
                })

            if dates_booked_col and dates:
                cell_val = fmt_dates_booked(dates)
                updates.append({
                    "range":  gspread.utils.rowcol_to_a1(row_idx, dates_booked_col),
                    "values": [[cell_val]]
                })

        if updates:
            ws.batch_update(updates)
            print(f"  [{tab}] {len(updates)} cells updated")
            updated += len(updates)
        else:
            print(f"  [{tab}] No matches")

    if no_data:
        print(f"\n  No budget data for {len(no_data)} models:")
        for n in sorted(no_data)[:15]:
            print(f"    {n}")
        if len(no_data) > 15:
            print(f"    ... and {len(no_data) - 15} more")

    return updated


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Calculate AVG Rates + Booking History")
    print("=" * 60)

    print(f"\nScanning: {BUDGET_DIR}")
    rate_hist, date_hist, canonical = build_histories()
    print(f"\nFound data for {len(rate_hist)} unique talent names\n")

    # Top 20 by booking count
    top = sorted(
        ((k, rate_hist[k], date_hist[k]) for k in rate_hist),
        key=lambda x: len(set(x[2])),
        reverse=True
    )[:20]
    print("Top 20 most-booked talent:")
    for key, rates, dates in top:
        display = canonical.get(key, key)
        print(f"  {display:<30} {rate_summary(rates, dates)}")

    print("\nUpdating Google Sheet...")
    gc      = get_gspread()
    count   = update_sheet(gc, rate_hist, date_hist, canonical)
    print(f"\nDone. {count} cells updated.")


if __name__ == "__main__":
    main()
