"""
Compliance router — iPad-based talent documentation workflow.

Handles the pre-production legal compliance package for BG shoots:
  1. Get BG shoots for a date with Drive folder / photo status
  2. Prepare Drive folder (create if missing, copy templates, fill dates)
  3. Upload ID + verification photos to Drive (and optionally MEGA)
  4. Sync completed package to MEGA Grail/{studio}/{scene_id}/Legal/

Routes:
  GET  /api/compliance/shoots?date=YYYY-MM-DD    — BG shoots for a date
  POST /api/compliance/shoots/{id}/prepare        — ensure Drive folder + PDFs
  POST /api/compliance/shoots/{id}/photos         — upload photos (multipart)
  POST /api/compliance/shoots/{id}/mega-sync      — copy everything to MEGA
"""

from __future__ import annotations

import asyncio
import io
import json
import logging
import shutil
import subprocess
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import base64
import re

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from pydantic import BaseModel, Field

from api.auth import CurrentUser
from api.config import get_settings
from api.auth import require_admin
from api.database import get_db
from api.compliance_db import (
    SignatureSearchHit,
    SignedTalent,
    W9Record,
    contract_version,
    get_signed_pdf_path,
    is_shoot_complete,
    list_signed_talents,
    list_w9_records,
    search_signatures,
    upsert_signature,
)
from api.compliance_photos_db import (
    StoredPhoto,
    count_by_shoot as _count_photos_by_shoot,
    delete_photo as _db_delete_photo,
    get_photo as _db_get_photo,
    list_photos as _db_list_photos,
    upsert_photo as _db_upsert_photo,
)
from api.compliance_pdf import render_agreement_pdf
from api.routers.shoots import (
    LEGAL_ROOT_FOLDER,
    _get_drive_rw_token,
    _get_drive_token,
    _load_shoots_window,
    _parse_shoot_date,
)

_log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/compliance", tags=["compliance"])

# ─── Drive template IDs (mirrored from legal_docs_run.mjs) ───────────────────

FEMALE_TPL_ID = "1ey06iXodjkOhK6BK-Q9nsQ2UtiaAo5Oc"
MALE_TPLS: dict[str, str] = {
    "MikeMancini":  "1xlbezjCXjTkxBwyI-QdaS4JGyCeESSc-",
    "JaydenMarcos": "13qNXBqrygLrcnMsKI9sZy_3XOYP0Gjok",
    "DannySteele":  "1PxyfZnZZqeOe4DiqNN7AmeQuNcrfXc2j",
}
DATE_FIELDS = {"Date 1", "Date 2", "Custom Field 13"}

# Drive folder ID for /Legal Docs 2026/IDS/. Holds the canonical male
# ID photos (front, back, bunny-ear verification). prepare_compliance
# copies the right slugs into each shoot folder so the human doesn't
# have to dig them up every shoot. Filenames in this folder are
# inconsistent (HEIC vs jpeg, hyphenated vs camel-cased, plural vs
# singular), so we register the exact source name per (talent, kind).
MALE_IDS_FOLDER = "1-TOsGQGGb3Klc38k_i4BFfUYb9jsYhC-"
MALE_IDS: dict[str, dict[str, str]] = {
    "DannySteele": {
        "front": "DannyFronts.HEIC",
        "back":  "DannyBacks.HEIC",
        "bunny": "DannyBunnyEars.HEIC",
    },
    "JaydenMarcos": {
        "front": "Jayden Fronts.HEIC",
        "back":  "Jayden Backs.HEIC",
        "bunny": "JaydenBunnyEars.HEIC",
    },
    "MikeMancini": {
        "front": "MikeMancini-IDs-Front.jpeg",
        # Mike's back is a PDF rather than an image — keep verbatim.
        "back":  "MikeMancini-ID-Back.pdf",
        "bunny": "MikeMancini-BunnyEars.jpeg",
    },
    "PierceParis": {
        # Pierce has multiple front/back exposures; default to the first.
        "front": "Pierce-Front1.jpeg",
        "back":  "Pierce-Back1.jpeg",
        "bunny": "PierceParis-BunnyEars.jpeg",
    },
}


# Scene types that the compliance flow handles. Originally just BG / BGCP
# (mixed boy/girl shoots that need both male + female paperwork), extended
# 2026-05-05 to include SOLO (VRAllure female-only scenes) and JOI
# (NaughtyJOI female-only scenes) — both produce shoot days for talent who
# still need W-9 / 2257 / agreement on file. The downstream prepare and sign
# flows already handle female-only by skipping male PDF copy when no male
# is set on the shoot.
COMPLIANCE_SCENE_TYPES = ("BG", "BGCP", "SOLO", "JOI")


def _ext_from_name(name: str) -> str:
    """Lowercase extension including the leading dot, or empty string."""
    if "." not in name:
        return ""
    return "." + name.rsplit(".", 1)[-1].lower()


def _copy_male_ids_to_shoot_folder(
    male_slug: str, dest_folder_id: str, existing_names: set[str], token: str,
) -> list[str]:
    """Copy each kind (front/back/bunny) of a male's IDs into the shoot folder.

    Renames the source to the conventional ``<MaleSlug>-id-front.<ext>``
    pattern so the validation in ``shoots.py`` picks it up automatically.
    Skips kinds already present (idempotent on re-prepare).

    Returns the list of newly-copied destination filenames.
    """
    spec = MALE_IDS.get(male_slug)
    if not spec:
        return []
    copied: list[str] = []
    DEST_FOR_KIND = {
        "front": "id-front",
        "back":  "id-back",
        "bunny": "bunny-ear",
    }
    # List the IDS folder once and index by name (lowercase) for fast lookup.
    ids_files = {f.get("name", "").lower(): f for f in _list_folder_files(MALE_IDS_FOLDER, token)}
    for kind, src_name in spec.items():
        src = ids_files.get(src_name.lower())
        if not src:
            _log.warning("MALE_IDS: %s/%s missing in /IDS/ (looked for %r)",
                         male_slug, kind, src_name)
            continue
        ext = _ext_from_name(src_name)
        dst_label = DEST_FOR_KIND.get(kind, kind)
        dst_name = f"{male_slug}-{dst_label}{ext}"
        if dst_name.lower() in existing_names:
            continue
        try:
            _copy_file(src["id"], dest_folder_id, dst_name, token)
            copied.append(dst_name)
        except Exception as exc:
            _log.warning("MALE_IDS copy %s -> %s failed: %s", src_name, dst_name, exc)
    return copied

# S4 mapping is centralized in s4_client._STUDIO_ALIASES — accept the legacy
# UI/Grail-tab names and resolve at upload time.
import s4_client


# ─── Response models ──────────────────────────────────────────────────────────

class ComplianceShoot(BaseModel):
    shoot_id: str
    shoot_date: str
    female_talent: str
    male_talent: str
    drive_folder_url: Optional[str] = None
    drive_folder_id: Optional[str] = None
    drive_folder_name: Optional[str] = None
    pdfs_ready: bool = False
    photos_uploaded: int = 0
    is_complete: bool = False
    scene_id: str = ""
    studio: str = ""


class PrepareResult(BaseModel):
    folder_id: str
    folder_url: str
    folder_name: str
    female_pdf_id: str = ""
    male_pdf_id: str = ""
    male_known: bool = False
    dates_filled: bool = False
    message: str = ""


class PhotoUploadResult(BaseModel):
    uploaded: list[str] = []
    drive_file_ids: list[str] = []
    mega_paths: list[str] = []
    errors: list[str] = []


class InitUploadFile(BaseModel):
    filename: str
    mime_type: str = "image/jpeg"


class UploadSession(BaseModel):
    filename: str
    upload_url: str


class ConfirmUploadsRequest(BaseModel):
    filenames: list[str]
    file_ids: list[str] = []


class MegaSyncResult(BaseModel):
    status: str  # ok | error
    mega_path: str = ""
    files_copied: int = 0
    message: str = ""


class FillFormRequest(BaseModel):
    talent: str          # "female" | male stage name e.g. "MikeMancini"
    legal_name: str
    stage_name: str = ""
    dob: str = ""        # YYYY-MM-DD from date input
    place_of_birth: str = ""
    street_address: str = ""
    city_state_zip: str = ""
    phone: str = ""
    email: str = ""
    id1_type: str = ""
    id1_number: str = ""
    id2_type: str = ""
    id2_number: str = ""
    signature: str = ""
    company_name: str = ""


# ─── Hub-only signing flow (TKT-0150) ────────────────────────────────────────


class SignRequest(BaseModel):
    """Single payload from the iPad form. Drives the new compliance flow:
    captures every field that lives on the legacy Drive PDF templates plus a
    drawn signature image (base64-encoded PNG)."""
    talent_role: str = Field(..., pattern=r"^(female|male)$")
    talent_slug: str
    talent_display: str

    # W-9 (page 1 of legacy template)
    legal_name: str
    business_name: str = ""
    tax_classification: str = Field("individual", pattern=r"^(individual|c_corp|s_corp|partnership|trust_estate|llc|other)$")
    llc_class: str = ""              # 'C' | 'S' | 'P' (only used when tax_classification='llc')
    other_classification: str = ""
    exempt_payee_code: str = ""
    fatca_code: str = ""
    tin_type: str = Field("ssn", pattern=r"^(ssn|ein)$")
    tin: str                          # raw digits, no formatting

    # 2257 Performer Names Disclosure (page 6 of legacy template)
    dob: str                          # YYYY-MM-DD
    place_of_birth: str
    street_address: str
    city_state_zip: str
    phone: str
    email: str
    id1_type: str
    id1_number: str
    id2_type: str = ""
    id2_number: str = ""
    stage_names: str = ""
    professional_names: str = ""
    nicknames_aliases: str = ""
    previous_legal_names: str = ""

    # Drawn signature, sent as a base64 data-URL or raw base64 PNG
    signature_png: str


class SignResult(BaseModel):
    shoot_id: str
    talent_role: str
    talent_slug: str
    signed_at: str
    pdf_local_path: str
    pdf_mega_path: str
    contract_version: str


class SignedSummary(BaseModel):
    """Per-talent summary returned by /shoots/{id}/signed for the UI."""
    talent_role: str
    talent_slug: str
    talent_display: str
    legal_name: str
    signed_at: str
    pdf_mega_path: str
    id: int = 0   # compliance_signatures.id — needed for the edit modal


# ─── Drive helpers ────────────────────────────────────────────────────────────


def _get_drive_oauth_token() -> Optional[str]:
    """
    Return a fresh access token derived from vr_oauth_token.json.

    The service account has 0 storage quota, so any file the SA *creates*
    in My Drive fails with 403 storageQuotaExceeded. Photo uploads must use
    the user's OAuth credentials so files are owned by — and consume the
    quota of — the configured Google user. Same pattern as call_sheets.py.
    """
    from api.config import get_settings
    try:
        token_path = get_settings().base_dir / "vr_oauth_token.json"
        if not token_path.exists():
            return None
        with open(token_path) as f:
            tok = json.load(f)
        params = urllib.parse.urlencode({
            "grant_type":    "refresh_token",
            "refresh_token": tok["refresh_token"],
            "client_id":     tok["client_id"],
            "client_secret": tok["client_secret"],
        }).encode()
        req = urllib.request.Request(
            tok.get("token_uri", "https://oauth2.googleapis.com/token"),
            data=params, method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read()).get("access_token")
    except Exception as exc:
        _log.warning("drive oauth token fetch failed: %s", exc)
        return None


def _drive_json(url: str, token: str, method: str = "GET",
                body: Optional[bytes] = None,
                content_type: str = "application/json") -> dict:
    headers: dict[str, str] = {"Authorization": f"Bearer {token}"}
    if body is not None:
        headers["Content-Type"] = content_type
    req = urllib.request.Request(url, data=body, method=method, headers=headers)
    with urllib.request.urlopen(req, timeout=20) as r:
        return json.loads(r.read())


def _find_or_create_folder(parent_id: str, name: str, token: str) -> str:
    q = (
        f"'{parent_id}' in parents and name={json.dumps(name)} "
        "and mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    res = _drive_json(
        "https://www.googleapis.com/drive/v3/files"
        f"?q={urllib.parse.quote(q)}&fields=files(id,name)",
        token,
    )
    existing = (res or {}).get("files") or []
    if existing:
        return existing[0]["id"]
    body = json.dumps({
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }).encode()
    created = _drive_json(
        "https://www.googleapis.com/drive/v3/files",
        token, method="POST", body=body,
    )
    return created["id"]


def _copy_file(file_id: str, dest_folder_id: str, new_name: str, token: str) -> dict:
    body = json.dumps({"name": new_name, "parents": [dest_folder_id]}).encode()
    return _drive_json(
        f"https://www.googleapis.com/drive/v3/files/{file_id}/copy",
        token, method="POST", body=body,
    )


def _list_folder_files(folder_id: str, token: str) -> list[dict]:
    q = f"'{folder_id}' in parents and trashed=false"
    res = _drive_json(
        "https://www.googleapis.com/drive/v3/files"
        f"?q={urllib.parse.quote(q)}&fields=files(id,name,mimeType,webViewLink)&pageSize=1000",
        token,
    )
    return (res or {}).get("files") or []


def _create_resumable_session(folder_id: str, file_name: str, mime_type: str, token: str) -> str:
    """
    Create a Drive resumable-upload session.
    Returns the session URI (upload URL).  The caller — typically the browser —
    can PUT the file body directly to this URL with no Authorization header.
    The URI is valid for ~7 days and is pre-authorized via the service account.
    """
    metadata = json.dumps({"name": file_name, "parents": [folder_id]}).encode()
    req = urllib.request.Request(
        "https://www.googleapis.com/upload/drive/v3/files"
        "?uploadType=resumable&fields=id",
        data=metadata,
        method="POST",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json; charset=UTF-8",
            "X-Upload-Content-Type": mime_type,
        },
    )
    with urllib.request.urlopen(req, timeout=20) as r:
        location = r.headers.get("Location", "")
    if not location:
        raise RuntimeError(f"Drive did not return a session URI for {file_name}")
    return location


def _upload_to_drive(folder_id: str, file_name: str,
                     file_bytes: bytes, mime_type: str, token: str) -> str:
    """Multipart upload to Drive. Returns new file ID.

    Raises with Drive's actual error reason on failure so the caller can
    surface it to the UI (otherwise urllib's default repr says only
    "HTTP Error 403: Forbidden", hiding the underlying cause).
    """
    boundary = b"ec_boundary_01"
    metadata = json.dumps({"name": file_name, "parents": [folder_id]}).encode()
    body = (
        b"--" + boundary + b"\r\n"
        b"Content-Type: application/json; charset=UTF-8\r\n\r\n"
        + metadata + b"\r\n"
        b"--" + boundary + b"\r\n"
        b"Content-Type: " + mime_type.encode() + b"\r\n\r\n"
        + file_bytes + b"\r\n"
        b"--" + boundary + b"--\r\n"
    )
    req = urllib.request.Request(
        "https://www.googleapis.com/upload/drive/v3/files"
        "?uploadType=multipart&fields=id&supportsAllDrives=true",
        data=body, method="POST",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": f"multipart/related; boundary={boundary.decode()}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            return json.loads(r.read()).get("id", "")
    except urllib.error.HTTPError as exc:
        try:
            err_body = exc.read().decode("utf-8", errors="replace")
            err_json = json.loads(err_body)
            reason = (err_json.get("error") or {}).get("message") or err_body[:300]
        except Exception:
            reason = f"HTTP {exc.code}"
        raise RuntimeError(f"Drive {exc.code}: {reason}") from exc


def _format_dob(dob_iso: str) -> str:
    try:
        d = datetime.strptime(dob_iso, "%Y-%m-%d")
        return d.strftime("%b ") + str(d.day) + d.strftime(", %Y")
    except Exception:
        return dob_iso


def _fill_pdf_fields(pdf_bytes: bytes, fields: dict[str, str]) -> bytes:
    """Fill arbitrary text/checkbox fields in a PDF."""
    try:
        import pypdf
        from pypdf.generic import NameObject, create_string_object
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        writer = pypdf.PdfWriter()
        writer.append(reader)
        for page in writer.pages:
            if "/Annots" not in page:
                continue
            for annot in page["/Annots"]:
                obj = annot.get_object()
                field_name = obj.get("/T")
                if field_name not in fields:
                    continue
                value = fields[field_name]
                field_type = str(obj.get("/FT", ""))
                if field_type == "/Btn":
                    name_val = NameObject(value) if value.startswith("/") else NameObject("/Off")
                    obj.update({
                        NameObject("/V"): name_val,
                        NameObject("/AS"): name_val,
                    })
                else:
                    obj.update({
                        NameObject("/V"):  create_string_object(value),
                        NameObject("/DV"): create_string_object(value),
                    })
        buf = io.BytesIO()
        writer.write(buf)
        return buf.getvalue()
    except Exception as exc:
        _log.warning("pdf fill failed: %s", exc)
        return pdf_bytes


def _row_to_pdf_fields(row: dict, shoot_date_str: str) -> dict[str, str]:
    """Build the PDF AcroForm field values from a compliance_signatures row.

    Mirror of `_map_form_to_pdf` for the DB-driven render path. Fields are
    typed as strings — date / phone / etc. formatting that the talent-form
    UI does is preserved verbatim because the row already stores the
    normalized values.
    """
    legal = row.get("legal_name", "") or ""
    biz = row.get("business_name", "") or ""
    stage = row.get("stage_names", "") or legal
    full_address = f"{row.get('street_address','') or ''} {row.get('city_state_zip','') or ''}".strip()
    dob_fmt = _format_dob(row.get("dob", "")) if row.get("dob") else ""
    return {
        "Custom Field 1":    legal,
        "Custom Field 2":    biz,
        "Custom Checkbox 1": "/Yes",
        "Custom Field 7":    row.get("street_address", "") or "",
        "Custom Field 9":    row.get("city_state_zip", "") or "",
        "Custom Field 14":   legal,
        "Custom Field 16":   legal,
        "Custom Field 17":   stage,
        "Custom Field 18":   legal,
        "Custom Field 23":   legal,
        "Custom Field 24":   legal,
        "Custom Field 25":   dob_fmt,
        "Custom Field 26":   row.get("place_of_birth", "") or "",
        "Custom Field 27":   full_address,
        "Custom Field 28":   row.get("id1_type", "") or "",
        "Custom Field 29":   row.get("id1_number", "") or "",
        "Custom Field 30":   row.get("id2_type", "") or "",
        "Custom Field 31":   row.get("id2_number", "") or "",
        "Custom Field 32":   row.get("phone", "") or "",
        "Custom Field 33":   row.get("email", "") or "",
        "Custom Field 34":   stage,
        "Custom Field 35":   stage,
        "Date 1":            shoot_date_str,
        "Date 2":            shoot_date_str,
        "Custom Field 13":   shoot_date_str,
    }


def _render_pdf_from_signature(row: dict, date_override: Optional[str] = None) -> bytes:
    """Render a filled-PDF from a compliance_signatures row.

    Uses the appropriate template (female: FEMALE_TPL_ID; male: MALE_TPLS
    keyed on slug). `date_override` lets the time-travel UI request a
    different date than what's stored on the row (e.g. "Apr 15, 2026"
    formatted via shoot_date.strftime("%b %-d, %Y")).
    """
    role = row.get("talent_role", "")
    slug = row.get("talent_slug", "")
    template_id: Optional[str] = None
    if role == "female":
        template_id = FEMALE_TPL_ID
    elif role == "male":
        template_id = MALE_TPLS.get(slug)
    if not template_id:
        raise HTTPException(status_code=400,
                            detail=f"No PDF template for role={role!r} slug={slug!r}")

    if date_override:
        date_str = date_override
    else:
        # Format YYYY-MM-DD as "Mon D, YYYY" — mirrors the existing flow.
        try:
            d = datetime.fromisoformat((row.get("shoot_date") or "")[:10])
            date_str = d.strftime("%b ") + str(d.day) + d.strftime(", %Y")
        except Exception:
            date_str = row.get("shoot_date", "") or ""

    token = _get_drive_token()
    if not token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")
    template_bytes = _drive_download_bytes(template_id, token)
    fields = _row_to_pdf_fields(row, date_str)
    return _fill_pdf_fields(template_bytes, fields)


def _map_form_to_pdf(req: "FillFormRequest", shoot_date_str: str) -> dict[str, str]:
    full_address = f"{req.street_address} {req.city_state_zip}".strip()
    dob_fmt = _format_dob(req.dob) if req.dob else ""
    return {
        "Custom Field 1":    req.legal_name,
        "Custom Field 2":    req.company_name,
        "Custom Checkbox 1": "/Yes",
        "Custom Field 7":    req.street_address,
        "Custom Field 9":    req.city_state_zip,
        "Custom Field 14":   req.legal_name,
        "Custom Field 16":   req.legal_name,
        "Custom Field 17":   req.stage_name,
        "Custom Field 18":   req.legal_name,
        "Custom Field 23":   req.legal_name,
        "Custom Field 24":   req.legal_name,
        "Custom Field 25":   dob_fmt,
        "Custom Field 26":   req.place_of_birth,
        "Custom Field 27":   full_address,
        "Custom Field 28":   req.id1_type,
        "Custom Field 29":   req.id1_number,
        "Custom Field 30":   req.id2_type,
        "Custom Field 31":   req.id2_number,
        "Custom Field 32":   req.phone,
        "Custom Field 33":   req.email,
        "Custom Field 34":   req.stage_name,
        "Custom Field 35":   req.stage_name,
        "Signature 1":       req.signature,
        "Signature 2":       req.signature,
        "Signature 3":       req.signature,
        "Date 1":            shoot_date_str,
        "Date 2":            shoot_date_str,
        "Custom Field 13":   shoot_date_str,
    }


def _fill_pdf_dates(pdf_bytes: bytes, today_str: str) -> bytes:
    try:
        import pypdf
        from pypdf.generic import NameObject, create_string_object
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        writer = pypdf.PdfWriter()
        writer.append(reader)
        for page in writer.pages:
            if "/Annots" not in page:
                continue
            for annot in page["/Annots"]:
                obj = annot.get_object()
                if obj.get("/T") in DATE_FIELDS:
                    obj.update({
                        NameObject("/V"):  create_string_object(today_str),
                        NameObject("/DV"): create_string_object(today_str),
                    })
        buf = io.BytesIO()
        writer.write(buf)
        return buf.getvalue()
    except Exception as exc:
        _log.warning("pdf date fill failed: %s", exc)
        return pdf_bytes


# In-memory TTL cache for shoot-folder lookups. The compliance list endpoint
# loops over a day's shoots and calls this for each — pre-cache the result
# so a second hit (page reload, modal reopen, polling) doesn't repeat the
# 2 sequential Drive REST calls. 5 min TTL covers a typical reviewer session.
_SHOOT_FOLDER_CACHE: dict[tuple[date, str, str], tuple[float, Optional[tuple[str, str]]]] = {}
_SHOOT_FOLDER_TTL_S = 300.0


def _get_shoot_folder(
    shoot_date: date, female: str, male: str, token: str
) -> Optional[tuple[str, str]]:
    """Return (folder_id, folder_name) if Drive folder already exists."""
    import time as _time

    cache_key = (shoot_date, (female or "").strip().lower(), (male or "").strip().lower())
    cached = _SHOOT_FOLDER_CACHE.get(cache_key)
    if cached and (_time.time() - cached[0]) < _SHOOT_FOLDER_TTL_S:
        return cached[1]

    month_name = shoot_date.strftime("%B")
    date_prefix = shoot_date.strftime("%m%d%y") + "-"
    female_slug = female.replace(" ", "").lower()
    male_slug = male.replace(" ", "").lower() if male else ""

    q = (
        f"'{LEGAL_ROOT_FOLDER}' in parents and name='{month_name}' "
        "and mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    res = _drive_json(
        "https://www.googleapis.com/drive/v3/files"
        f"?q={urllib.parse.quote(q)}&fields=files(id,name)",
        token,
    )
    month_files = (res or {}).get("files") or []
    if not month_files:
        _SHOOT_FOLDER_CACHE[cache_key] = (_time.time(), None)
        return None
    month_id = month_files[0]["id"]

    q2 = (
        f"'{month_id}' in parents "
        "and mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    res2 = _drive_json(
        "https://www.googleapis.com/drive/v3/files"
        f"?q={urllib.parse.quote(q2)}&fields=files(id,name)&pageSize=1000",
        token,
    )
    for f in (res2 or {}).get("files") or []:
        fname: str = f.get("name", "")
        if not fname.startswith(date_prefix):
            continue
        name_lower = fname.replace("-", "").lower()
        if female_slug and female_slug not in name_lower:
            continue
        if male_slug and male_slug not in name_lower:
            continue
        result = (f["id"], fname)
        _SHOOT_FOLDER_CACHE[cache_key] = (_time.time(), result)
        return result
    _SHOOT_FOLDER_CACHE[cache_key] = (_time.time(), None)
    return None


def _window() -> tuple[date, date]:
    today = datetime.now(timezone.utc).date()
    return today - timedelta(days=60), today + timedelta(days=30)


# ─── Routes ──────────────────────────────────────────────────────────────────

@router.get("/shoots", response_model=list[ComplianceShoot])
async def list_compliance_shoots(
    user: CurrentUser,
    date: Optional[str] = Query(default=None),
    q: Optional[str] = Query(default=None, description="Search by talent name across a wide date window"),
):
    """Return BG shoots for a given date (default: today) with compliance status.

    When `q` is provided, the date filter is widened to a year of context and
    shoots are filtered to those whose female or male talent name matches the
    query (case-insensitive substring). Used by the compliance page's name
    search so admins can find a specific talent's shoot without paging by date.
    """
    if q and q.strip():
        # Wide window: 1 year back, 4 months forward — enough for tax-year work
        today = datetime.now(timezone.utc).date()
        window_lo = today - timedelta(days=365)
        window_hi = today + timedelta(days=120)
        shoots = _load_shoots_window(window_lo, window_hi, include_cancelled=False)
        needle = q.strip().lower()
        shoots = [
            s for s in shoots
            if needle in s.female_talent.lower()
            or needle in (s.male_talent or "").lower()
        ]
    else:
        target_date_str = date or datetime.now(timezone.utc).date().isoformat()
        target_date = _parse_shoot_date(target_date_str)
        if target_date is None:
            raise HTTPException(status_code=400, detail="Invalid date")
        shoots = _load_shoots_window(target_date, target_date, include_cancelled=False)
    token = _get_drive_token()
    # Bulk-fetch DB-backed signatures so is_complete + pdfs_ready don't depend
    # on Drive folder presence (the Drive proxy false-positives the moment
    # prepare copies blank templates — see TKT-0150).
    shoot_ids = [s.shoot_id for s in shoots]
    signed_by_shoot = list_signed_talents(shoot_ids)
    photo_counts_db = _count_photos_by_shoot(shoot_ids)
    results: list[ComplianceShoot] = []

    for shoot in shoots:
        bg_scenes = [sc for sc in shoot.scenes if sc.scene_type.upper() in COMPLIANCE_SCENE_TYPES]
        if not bg_scenes:
            continue

        primary = bg_scenes[0]
        scene_id = primary.scene_id or ""
        studio = primary.studio or ""

        folder_url = folder_id = folder_name = None
        photos_count = 0

        # Drive folder lookup is now PURELY for displaying the legacy folder
        # link in the UI. Completion comes from compliance_signatures, and
        # the canonical photo count is `photo_counts_db` (server-persisted).
        #
        # We avoid the per-shoot `_list_folder_files` call when DB already
        # has a count — that's the slowest leg of this loop (~500ms/shoot
        # cold) and was the main reason this endpoint took double-digit
        # seconds with a busy day.
        db_photo_count = photo_counts_db.get(shoot.shoot_id, 0)
        if token:
            shoot_date_obj = _parse_shoot_date(shoot.shoot_date)
            if shoot_date_obj:
                try:
                    info = _get_shoot_folder(
                        shoot_date_obj, shoot.female_talent, shoot.male_talent, token
                    )
                    if info:
                        folder_id, folder_name = info
                        folder_url = f"https://drive.google.com/drive/folders/{folder_id}"
                        # Only walk Drive when DB doesn't have the count.
                        # New uploads land in DB synchronously, so the
                        # fallback is just for legacy/back-filled shoots.
                        if db_photo_count == 0:
                            files = _list_folder_files(folder_id, token)
                            photos_count = sum(
                                1 for f in files
                                if f.get("name", "").lower().endswith((".jpg", ".jpeg", ".png"))
                            )
                except Exception as exc:
                    _log.debug("compliance folder lookup: %s", exc)

        # Completion: prefer DB-backed signatures (new flow), fall back to the
        # legacy Drive-folder check so shoots completed via the old prepare/
        # fill-form path still show as complete during the cutover. Once the
        # Hub UI is fully rewired, the Drive fallback can be removed.
        signed = signed_by_shoot.get(shoot.shoot_id, [])
        signed_roles = {t.talent_role for t in signed}
        needed_roles = {"female", "male"} if shoot.male_talent else {"female"}
        if signed:
            is_complete = needed_roles.issubset(signed_roles)
            pdfs_ready = True
        else:
            # Legacy fallback — mirror the pre-TKT-0150 logic
            is_complete_legacy = False
            for sc in bg_scenes:
                for asset in sc.assets:
                    if asset.asset_type == "legal_docs_uploaded" and asset.status == "validated":
                        is_complete_legacy = True
            is_complete = is_complete_legacy
            # pdfs_ready ← legacy file-count check
            need_pdfs = 2 if shoot.male_talent else 1
            try:
                if folder_id and token:
                    files_for_pdf = _list_folder_files(folder_id, token)
                    pdfs_ready = sum(
                        1 for f in files_for_pdf
                        if f.get("name", "").lower().endswith(".pdf")
                    ) >= need_pdfs
                else:
                    pdfs_ready = False
            except Exception:
                pdfs_ready = False

        # Photo count: prefer DB-backed (server-side persisted) row count;
        # fall back to whatever the Drive lookup found so existing-folder
        # shoots from the legacy flow still display a non-zero count.
        photos_total = max(photo_counts_db.get(shoot.shoot_id, 0), photos_count)

        results.append(ComplianceShoot(
            shoot_id=shoot.shoot_id,
            shoot_date=shoot.shoot_date,
            female_talent=shoot.female_talent,
            male_talent=shoot.male_talent or "",
            drive_folder_url=folder_url,
            drive_folder_id=folder_id,
            drive_folder_name=folder_name,
            pdfs_ready=pdfs_ready,
            photos_uploaded=photos_total,
            is_complete=is_complete,
            scene_id=scene_id,
            studio=studio,
        ))

    return results


@router.post("/shoots/{shoot_id}/prepare", response_model=PrepareResult)
async def prepare_compliance(shoot_id: str, user: CurrentUser):
    """
    Ensure Drive legal folder exists, copy PDF templates, and fill
    date fields in the male PDF.
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    shoot_date = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date:
        raise HTTPException(status_code=400, detail="Invalid shoot date")

    female = shoot.female_talent.strip()
    male = (shoot.male_talent or "").strip()

    token = _get_drive_rw_token()
    if not token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")

    # 1. Month folder
    month_name = shoot_date.strftime("%B")
    month_id = _find_or_create_folder(LEGAL_ROOT_FOLDER, month_name, token)

    # 2. Shoot folder
    date_code = shoot_date.strftime("%m%d%y")
    female_slug = female.replace(" ", "")
    male_slug = male.replace(" ", "")
    folder_name = f"{date_code}-{female_slug}" + (f"-{male_slug}" if male_slug else "")
    folder_id = _find_or_create_folder(month_id, folder_name, token)
    folder_url = f"https://drive.google.com/drive/folders/{folder_id}"

    # 3. Existing files
    existing = _list_folder_files(folder_id, token)
    existing_lower = {f.get("name", "").lower() for f in existing}

    female_pdf_name = f"{female_slug}-{date_code}.pdf"
    male_pdf_name = f"{male_slug}-{date_code}.pdf" if male_slug else ""
    female_pdf_id = male_pdf_id = ""
    male_known = bool(male_slug and (male_slug in MALE_TPLS or male in MALE_TPLS))
    dates_filled = False

    # 4. Female PDF
    if female_pdf_name.lower() not in existing_lower:
        try:
            r = _copy_file(FEMALE_TPL_ID, folder_id, female_pdf_name, token)
            female_pdf_id = r.get("id", "")
        except Exception as exc:
            _log.warning("female PDF copy failed: %s", exc)
    else:
        for f in existing:
            if f.get("name", "").lower() == female_pdf_name.lower():
                female_pdf_id = f.get("id", "")
                break

    # 5. Male PDF (known templates only)
    if male_slug and male_pdf_name:
        male_tpl = MALE_TPLS.get(male_slug) or MALE_TPLS.get(male)
        if male_pdf_name.lower() not in existing_lower and male_tpl:
            try:
                r = _copy_file(male_tpl, folder_id, male_pdf_name, token)
                male_pdf_id = r.get("id", "")
            except Exception as exc:
                _log.warning("male PDF copy failed: %s", exc)
        elif male_pdf_name.lower() in existing_lower:
            for f in existing:
                if f.get("name", "").lower() == male_pdf_name.lower():
                    male_pdf_id = f.get("id", "")
                    break

    # 6. Fill dates in male PDF
    if male_pdf_id:
        try:
            today_str = shoot_date.strftime("%b %-d, %Y")
            dl = urllib.request.Request(
                f"https://www.googleapis.com/drive/v3/files/{male_pdf_id}?alt=media",
                headers={"Authorization": f"Bearer {token}"},
            )
            with urllib.request.urlopen(dl, timeout=20) as resp:
                pdf_bytes = resp.read()
            filled = _fill_pdf_dates(pdf_bytes, today_str)
            patch = urllib.request.Request(
                f"https://www.googleapis.com/upload/drive/v3/files/{male_pdf_id}?uploadType=media",
                data=filled, method="PATCH",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/pdf"},
            )
            with urllib.request.urlopen(patch, timeout=30) as r:
                r.read()
            dates_filled = True
        except Exception as exc:
            _log.warning("male date fill failed: %s", exc)

    # 7. Auto-copy male IDs from /Legal Docs 2026/IDS/ into the shoot folder.
    # Idempotent — skips any kind whose dest filename is already present.
    male_ids_copied: list[str] = []
    if male_slug and male_known:
        try:
            male_ids_copied = _copy_male_ids_to_shoot_folder(
                male_slug, folder_id, existing_lower, token,
            )
        except Exception as exc:
            _log.warning("male ID auto-copy failed: %s", exc)

    parts = []
    if female_pdf_id:
        parts.append(f"{female_slug} PDF ready")
    if male_pdf_id:
        parts.append(f"{male_slug} PDF {'+ dates ' if dates_filled else ''}ready")
    elif male_slug and not male_known:
        parts.append(f"⚠ {male_slug} not on file — upload manually")
    if male_ids_copied:
        parts.append(f"{male_slug} IDs ({len(male_ids_copied)}) copied")

    return PrepareResult(
        folder_id=folder_id,
        folder_url=folder_url,
        folder_name=folder_name,
        female_pdf_id=female_pdf_id,
        male_pdf_id=male_pdf_id,
        male_known=male_known,
        dates_filled=dates_filled,
        message="; ".join(parts) or "Folder ready",
    )


@router.post("/shoots/{shoot_id}/photos", response_model=PhotoUploadResult)
async def upload_photos(
    shoot_id: str,
    user: CurrentUser,
    files: list[UploadFile] = File(...),
    labels: list[str] = Form(default=[]),
    scene_id: str = Form(default=""),
    studio: str = Form(default=""),
):
    """
    Upload ID / verification photos to the Drive legal folder.
    Also copies to MEGA Legal/ subfolder if scene_id + studio are provided.
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    shoot_date = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date:
        raise HTTPException(status_code=400, detail="Invalid shoot date")

    # Photo uploads use the user's OAuth token so files are owned by — and
    # consume the quota of — the configured Google user. Service accounts
    # have 0 storage, so SA-owned multipart uploads to My Drive folders fail
    # with 403 storageQuotaExceeded. Folder lookup still works with either
    # token, so we fall back to the SA token if vr_oauth_token.json is missing.
    upload_token = _get_drive_oauth_token()
    lookup_token = upload_token or _get_drive_rw_token()
    if not lookup_token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")
    if not upload_token:
        upload_token = lookup_token

    folder_info = _get_shoot_folder(
        shoot_date, shoot.female_talent, shoot.male_talent, lookup_token
    )
    if not folder_info:
        raise HTTPException(
            status_code=404,
            detail="Drive folder not found — tap Prepare Docs first",
        )
    folder_id, _ = folder_info

    # ── Build (label, content, mime) tuples for every file ───────────────
    # Read all uploads concurrently — multipart bodies are already buffered
    # by Starlette so these awaits are fast memory reads, not network IO.
    async def _read_one(i: int, upload_file: UploadFile):
        label = labels[i] if i < len(labels) else (upload_file.filename or f"photo_{i + 1}.jpg")
        if not any(label.lower().endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".mp4", ".mov", ".webm")):
            label += ".jpg"
        content = await upload_file.read()
        mime = upload_file.content_type or "image/jpeg"
        return label, content, mime

    file_tuples = await asyncio.gather(*[_read_one(i, f) for i, f in enumerate(files)])

    # ── Upload to Drive in parallel (thread pool — Drive HTTP is blocking) ─
    async def _drive_one(label: str, content: bytes, mime: str):
        return await asyncio.to_thread(_upload_to_drive, folder_id, label, content, mime, upload_token)

    drive_results = await asyncio.gather(
        *[_drive_one(label, content, mime) for label, content, mime in file_tuples],
        return_exceptions=True,
    )

    uploaded: list[str] = []
    drive_file_ids: list[str] = []
    mega_paths: list[str] = []
    errors: list[str] = []

    for (label, content, _mime), result in zip(file_tuples, drive_results):
        if isinstance(result, Exception):
            _log.warning("photo upload failed %s: %s", label, result)
            errors.append(f"{label}: {result}")
        else:
            uploaded.append(label)
            drive_file_ids.append(result)

    # ── MEGA sync (single rclone call, only when caller requests it) ──────
    if scene_id and studio and uploaded:
        tmp_dir = tempfile.mkdtemp(prefix="compliance_")
        try:
            for (label, content, _mime) in file_tuples:
                if label in uploaded:
                    (Path(tmp_dir) / label).write_bytes(content)
            try:
                bucket = s4_client.STUDIO_BUCKETS[s4_client._STUDIO_ALIASES.get(studio, studio).upper()]
                for label in uploaded:
                    key = s4_client.key_for(scene_id, "Legal", label)
                    src = Path(tmp_dir) / label
                    await asyncio.to_thread(s4_client.put_object, studio, key, src)
                    mega_paths.append(f"s3://{bucket}/{key}")
            except Exception as exc:
                errors.append(f"S4: {exc}")
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    return PhotoUploadResult(
        uploaded=uploaded,
        drive_file_ids=drive_file_ids,
        mega_paths=mega_paths,
        errors=errors,
    )


@router.post("/shoots/{shoot_id}/mega-sync", response_model=MegaSyncResult)
async def mega_sync(
    shoot_id: str,
    user: CurrentUser,
    scene_id: str = Form(...),
    studio: str = Form(...),
):
    """
    Download every file from the Drive legal folder and push to
    MEGA Grail/{studio}/{scene_id}/Legal/.
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    shoot_date = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date:
        raise HTTPException(status_code=400, detail="Invalid shoot date")

    token = _get_drive_rw_token()
    if not token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")

    folder_info = _get_shoot_folder(
        shoot_date, shoot.female_talent, shoot.male_talent, token
    )
    if not folder_info:
        raise HTTPException(status_code=404, detail="Drive folder not found")
    folder_id, _ = folder_info

    tmp_dir = tempfile.mkdtemp(prefix="compliance_sync_")
    try:
        files = _list_folder_files(folder_id, token)
        downloaded = 0
        for f in files:
            name = f.get("name", "")
            fid = f.get("id", "")
            if not name or not fid:
                continue
            try:
                dl = urllib.request.Request(
                    f"https://www.googleapis.com/drive/v3/files/{fid}?alt=media",
                    headers={"Authorization": f"Bearer {token}"},
                )
                with urllib.request.urlopen(dl, timeout=30) as resp:
                    (Path(tmp_dir) / name).write_bytes(resp.read())
                downloaded += 1
            except Exception as exc:
                _log.warning("sync download failed %s: %s", name, exc)

        bucket = s4_client.STUDIO_BUCKETS[s4_client._STUDIO_ALIASES.get(studio, studio).upper()]
        s4_prefix = s4_client.key_for(scene_id, "Legal") + "/"
        try:
            uploaded = 0
            for f in Path(tmp_dir).iterdir():
                if not f.is_file():
                    continue
                key = s4_prefix + f.name
                s4_client.put_object(studio, key, f)
                uploaded += 1
            return MegaSyncResult(
                status="ok",
                mega_path=f"s3://{bucket}/{s4_prefix}",
                files_copied=uploaded,
                message=f"Copied {uploaded} file(s) to S4",
            )
        except Exception as exc:
            return MegaSyncResult(
                status="error",
                mega_path=f"s3://{bucket}/{s4_prefix}",
                message=str(exc)[:300],
            )
    except Exception as exc:
        return MegaSyncResult(status="error", message=str(exc))
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.get("/shoots/{shoot_id}/pdf")
async def get_filled_pdf(
    shoot_id: str,
    user: CurrentUser,
    talent: str = Query(...),
):
    """Serve the agreement PDF for a talent.

    Preferred source: `pdf_local_path` from `compliance_signatures` (written
    by the new /sign flow). Falls back to the legacy Drive folder so the
    existing prepare/fill-form path keeps working until the Hub UI is
    rewired to the new endpoints."""
    from fastapi.responses import Response as FastAPIResponse

    talent_slug = talent.replace(" ", "")

    # 1. New flow — DB-backed local PDF
    pdf_path = (
        get_signed_pdf_path(shoot_id, "female", talent_slug)
        or get_signed_pdf_path(shoot_id, "male", talent_slug)
    )
    if pdf_path:
        p = Path(pdf_path)
        if p.exists():
            return FastAPIResponse(
                content=p.read_bytes(),
                media_type="application/pdf",
                headers={"Content-Disposition": f"inline; filename={p.name}"},
            )

    # 2. Legacy fallback — Drive folder (the old fill-form path writes here)
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    shoot_date = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date:
        raise HTTPException(status_code=400, detail="Invalid shoot date")

    token = _get_drive_token()
    if not token:
        raise HTTPException(status_code=404, detail="Talent has not signed yet")

    folder_info = _get_shoot_folder(
        shoot_date, shoot.female_talent, shoot.male_talent, token
    )
    if not folder_info:
        raise HTTPException(status_code=404, detail="Talent has not signed yet")
    folder_id, _ = folder_info
    files = _list_folder_files(folder_id, token)
    date_code = shoot_date.strftime("%m%d%y")
    pdf_name = f"{talent_slug}-{date_code}.pdf"
    pdf_file = next(
        (f for f in files if f.get("name", "").lower() == pdf_name.lower()), None
    )
    if not pdf_file:
        raise HTTPException(status_code=404, detail=f"PDF not found: {pdf_name}")

    dl_req = urllib.request.Request(
        f"https://www.googleapis.com/drive/v3/files/{pdf_file['id']}?alt=media",
        headers={"Authorization": f"Bearer {token}"},
    )
    with urllib.request.urlopen(dl_req, timeout=30) as resp:
        pdf_bytes = resp.read()
    return FastAPIResponse(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={pdf_name}"},
    )


@router.post("/shoots/{shoot_id}/fill-form", response_model=PrepareResult)
async def fill_form(shoot_id: str, user: CurrentUser, req: FillFormRequest):
    """
    Generate a filled PDF from hub form data and save it to the Drive legal folder.
    Works for female talent (uses female template) and unknown male talent.
    Known male talent should use the prepare endpoint instead.
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    shoot_date = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date:
        raise HTTPException(status_code=400, detail="Invalid shoot date")

    token = _get_drive_rw_token()
    if not token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")

    # Select template: known males use their stored template; everyone else uses female template
    if req.talent == "female":
        tpl_id = FEMALE_TPL_ID
        talent_name = shoot.female_talent.strip()
    else:
        tpl_id = MALE_TPLS.get(req.talent) or FEMALE_TPL_ID
        talent_name = req.talent

    # Ensure Drive folder exists
    month_name = shoot_date.strftime("%B")
    month_id = _find_or_create_folder(LEGAL_ROOT_FOLDER, month_name, token)
    date_code = shoot_date.strftime("%m%d%y")
    female_slug = shoot.female_talent.strip().replace(" ", "")
    male_slug = (shoot.male_talent or "").strip().replace(" ", "")
    folder_name = f"{date_code}-{female_slug}" + (f"-{male_slug}" if male_slug else "")
    folder_id = _find_or_create_folder(month_id, folder_name, token)
    folder_url = f"https://drive.google.com/drive/folders/{folder_id}"

    # Download template
    dl_req = urllib.request.Request(
        f"https://www.googleapis.com/drive/v3/files/{tpl_id}?alt=media",
        headers={"Authorization": f"Bearer {token}"},
    )
    with urllib.request.urlopen(dl_req, timeout=30) as resp:
        pdf_bytes = resp.read()

    # Fill all fields
    date_str = shoot_date.strftime("%b ") + str(shoot_date.day) + shoot_date.strftime(", %Y")
    field_values = _map_form_to_pdf(req, date_str)
    filled = _fill_pdf_fields(pdf_bytes, field_values)

    # Upload or update in Drive folder
    talent_slug = talent_name.replace(" ", "")
    pdf_name = f"{talent_slug}-{date_code}.pdf"
    existing = _list_folder_files(folder_id, token)
    existing_file = next(
        (f for f in existing if f.get("name", "").lower() == pdf_name.lower()), None
    )

    if existing_file:
        patch = urllib.request.Request(
            f"https://www.googleapis.com/upload/drive/v3/files/{existing_file['id']}?uploadType=media",
            data=filled, method="PATCH",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/pdf"},
        )
        with urllib.request.urlopen(patch, timeout=30) as r:
            r.read()
        pdf_id = existing_file["id"]
    else:
        pdf_id = _upload_to_drive(folder_id, pdf_name, filled, "application/pdf", token)

    is_female = req.talent == "female"
    return PrepareResult(
        folder_id=folder_id,
        folder_url=folder_url,
        folder_name=folder_name,
        female_pdf_id=pdf_id if is_female else "",
        male_pdf_id=pdf_id if not is_female else "",
        male_known=req.talent in MALE_TPLS,
        dates_filled=True,
        message=f"{talent_slug} PDF saved to Drive",
    )


# ─── Direct-to-Drive upload endpoints ────────────────────────────────────────
# Instead of routing file bytes through this server, the browser gets a
# pre-authorized Drive resumable-session URL and uploads directly to Google.
# This eliminates the double-hop: Browser→Server→Drive becomes Browser→Drive.

@router.post("/shoots/{shoot_id}/photos/init-uploads", response_model=list[UploadSession])
async def init_photo_uploads(
    shoot_id: str,
    user: CurrentUser,
    files: list[InitUploadFile],
):
    """
    Create Drive resumable-upload sessions for a list of files.
    Returns one upload_url per file; the browser PUTs each file directly
    to that URL (no Authorization header needed — session is pre-authorized).
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    shoot_date = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date:
        raise HTTPException(status_code=400, detail="Invalid shoot date")

    token = _get_drive_rw_token()
    if not token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")

    folder_info = _get_shoot_folder(
        shoot_date, shoot.female_talent, shoot.male_talent, token
    )
    if not folder_info:
        raise HTTPException(
            status_code=404,
            detail="Drive folder not found — tap Prepare Docs first",
        )
    folder_id, _ = folder_info

    # Create all sessions concurrently (each is a small metadata-only request)
    async def _init_one(f: InitUploadFile) -> UploadSession:
        filename = f.filename
        if not any(filename.lower().endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".mp4", ".mov", ".webm")):
            filename += ".jpg"
        upload_url = await asyncio.to_thread(
            _create_resumable_session, folder_id, filename, f.mime_type, token
        )
        return UploadSession(filename=filename, upload_url=upload_url)

    sessions = await asyncio.gather(*[_init_one(f) for f in files], return_exceptions=True)

    results: list[UploadSession] = []
    for f, session in zip(files, sessions):
        if isinstance(session, Exception):
            _log.error("init upload session failed %s: %s", f.filename, session)
            raise HTTPException(status_code=502, detail=f"Could not create session for {f.filename}: {session}")
        results.append(session)  # type: ignore[arg-type]
    return results


@router.post("/shoots/{shoot_id}/photos/confirm-uploads", response_model=PhotoUploadResult)
async def confirm_photo_uploads(
    shoot_id: str,
    user: CurrentUser,
    body: ConfirmUploadsRequest,
):
    """
    Called by the browser after direct Drive uploads complete.
    Records the uploaded filenames so photos_uploaded count stays accurate.
    """
    # Validation: shoot must exist
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    return PhotoUploadResult(
        uploaded=body.filenames,
        drive_file_ids=body.file_ids,
        mega_paths=[],
        errors=[],
    )


# ─── Hub-only signing flow (TKT-0150) ────────────────────────────────────────
# These endpoints replace prepare/fill-form. The Hub renders the contract
# verbatim, captures a drawn signature, posts here. We persist to
# compliance_signatures, render our own PDF, push it to MEGA (no Drive).


_SIG_DATA_URL_RE = re.compile(r"^data:image/png;base64,(.+)$", re.IGNORECASE)


def _decode_signature_png(payload: str) -> bytes:
    """Accept either a data: URL or raw base64 PNG, return the PNG bytes."""
    if not payload:
        raise HTTPException(status_code=400, detail="signature_png required")
    m = _SIG_DATA_URL_RE.match(payload)
    raw_b64 = m.group(1) if m else payload
    try:
        b = base64.b64decode(raw_b64, validate=True)
    except Exception:
        raise HTTPException(status_code=400, detail="signature_png is not valid base64")
    if not b.startswith(b"\x89PNG\r\n\x1a\n"):
        raise HTTPException(status_code=400, detail="signature_png must be a PNG")
    return b


def _signature_dir() -> Path:
    """Local on-disk storage for raw signature PNGs (audit trail)."""
    p = Path(get_settings().base_dir) / "compliance_signatures"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _legal_pdf_dir() -> Path:
    """Local on-disk storage for generated agreement PDFs (pre-MEGA push)."""
    p = Path(get_settings().base_dir) / "compliance_pdfs"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _push_to_mega(local: Path, studio: str, key: str) -> Optional[str]:
    """Upload `local` to S4 at studio/key; returns error string on failure, None on success.

    Function name is preserved (rather than _push_to_s4) so the four legacy
    callsites read straightforwardly. Studio identifier is resolved by
    s4_client to its bucket; key is bucket-rooted (e.g. "VRH0030/Legal/x.pdf").
    """
    try:
        s4_client.put_object(studio, key, local)
        return None
    except Exception as exc:
        return str(exc)[:300]


@router.post("/shoots/{shoot_id}/sign", response_model=SignResult)
async def sign_shoot(shoot_id: str, user: CurrentUser, request: Request, body: SignRequest):
    """
    End-to-end Hub signing flow:
      1. Validate shoot + talent role
      2. Save the drawn signature PNG to disk (audit trail)
      3. Render our own agreement PDF embedding the signature
      4. Persist every field + audit metadata to compliance_signatures
      5. Push the PDF to MEGA Grail/{Studio}/{scene_id}/Legal/
      6. Return paths so the UI can link to the saved PDF
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    # Pick the BG scene we'll attach this signature to
    bg_scenes = [sc for sc in shoot.scenes if sc.scene_type.upper() in COMPLIANCE_SCENE_TYPES]
    if not bg_scenes:
        raise HTTPException(status_code=400, detail="Shoot has no BG/BGCP scene")
    primary = bg_scenes[0]
    scene_id = primary.scene_id or ""
    studio   = primary.studio or ""

    # Cross-check talent identity against the shoot record
    if body.talent_role == "female":
        expected = shoot.female_talent.replace(" ", "")
    else:
        expected = (shoot.male_talent or "").replace(" ", "")
    if not expected:
        raise HTTPException(status_code=400, detail=f"Shoot has no {body.talent_role} talent")
    if body.talent_slug.replace(" ", "").lower() != expected.lower():
        raise HTTPException(
            status_code=400,
            detail=f"talent_slug {body.talent_slug!r} does not match shoot's {body.talent_role}",
        )

    # 1. Save signature image to disk
    png_bytes = _decode_signature_png(body.signature_png)
    sig_path = _signature_dir() / f"{shoot.shoot_date}-{body.talent_slug}-{body.talent_role}.png"
    sig_path.write_bytes(png_bytes)

    # 2. Render PDF
    shoot_date_obj = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date_obj:
        raise HTTPException(status_code=400, detail="Invalid shoot date")
    signed_at_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    pdf_name = f"{body.talent_slug}-{shoot_date_obj.strftime('%m%d%y')}.pdf"
    pdf_path = _legal_pdf_dir() / shoot.shoot_date / pdf_name
    render_agreement_pdf(
        talent_display=body.talent_display,
        talent_role=body.talent_role,
        legal_name=body.legal_name,
        business_name=body.business_name,
        tax_classification=body.tax_classification,
        llc_class=body.llc_class,
        other_classification=body.other_classification,
        exempt_payee_code=body.exempt_payee_code,
        fatca_code=body.fatca_code,
        tin_type=body.tin_type,
        tin=body.tin,
        dob=body.dob,
        place_of_birth=body.place_of_birth,
        street_address=body.street_address,
        city_state_zip=body.city_state_zip,
        phone=body.phone,
        email=body.email,
        id1_type=body.id1_type,
        id1_number=body.id1_number,
        id2_type=body.id2_type,
        id2_number=body.id2_number,
        stage_names=body.stage_names,
        professional_names=body.professional_names,
        nicknames_aliases=body.nicknames_aliases,
        previous_legal_names=body.previous_legal_names,
        signature_png_bytes=png_bytes,
        shoot_date=shoot_date_obj,
        signed_at_iso=signed_at_iso,
        output_path=pdf_path,
    )

    # 3. Push to MEGA S4
    mega_remote = ""
    mega_err: Optional[str] = None
    if scene_id:
        s4_key = s4_client.key_for(scene_id, "Legal", pdf_name)
        bucket = s4_client.STUDIO_BUCKETS[s4_client._STUDIO_ALIASES.get(studio, studio).upper()]
        mega_remote = f"s3://{bucket}/{s4_key}"
        mega_err = _push_to_mega(pdf_path, studio, s4_key)
        if mega_err:
            _log.warning("S4 push failed for %s: %s", pdf_name, mega_err)
    else:
        _log.warning("Skipping S4 push: no scene_id for %s", shoot_id)

    # 4. Persist
    upsert_signature(
        shoot_id=shoot.shoot_id,
        shoot_date=shoot.shoot_date,
        scene_id=scene_id,
        studio=studio,
        talent_role=body.talent_role,
        talent_slug=body.talent_slug,
        talent_display=body.talent_display,
        legal_name=body.legal_name,
        business_name=body.business_name,
        tax_classification=body.tax_classification,
        llc_class=body.llc_class,
        other_classification=body.other_classification,
        exempt_payee_code=body.exempt_payee_code,
        fatca_code=body.fatca_code,
        tin_type=body.tin_type,
        tin=body.tin,
        dob=body.dob,
        place_of_birth=body.place_of_birth,
        street_address=body.street_address,
        city_state_zip=body.city_state_zip,
        phone=body.phone,
        email=body.email,
        id1_type=body.id1_type,
        id1_number=body.id1_number,
        id2_type=body.id2_type,
        id2_number=body.id2_number,
        stage_names=body.stage_names,
        professional_names=body.professional_names,
        nicknames_aliases=body.nicknames_aliases,
        previous_legal_names=body.previous_legal_names,
        signature_image_path=str(sig_path.relative_to(get_settings().base_dir)),
        signed_ip=request.client.host if request.client else "",
        signed_user_agent=request.headers.get("user-agent", "")[:300],
        signed_by_user=getattr(user, "email", "") or "",
        pdf_local_path=str(pdf_path),
        pdf_mega_path=mega_remote if not mega_err else "",
    )

    return SignResult(
        shoot_id=shoot.shoot_id,
        talent_role=body.talent_role,
        talent_slug=body.talent_slug,
        signed_at=signed_at_iso,
        pdf_local_path=str(pdf_path),
        pdf_mega_path=mega_remote if not mega_err else "",
        contract_version=contract_version(),
    )


@router.get("/shoots/{shoot_id}/signed", response_model=list[SignedSummary])
async def get_signed_summary(shoot_id: str, user: CurrentUser):
    """Return one row per talent who has completed the in-Hub agreement flow."""
    by_shoot = list_signed_talents([shoot_id])
    talents = by_shoot.get(shoot_id, [])
    return [
        SignedSummary(
            id=t.id,
            talent_role=t.talent_role,
            talent_slug=t.talent_slug,
            talent_display=t.talent_display,
            legal_name=t.legal_name,
            signed_at=t.signed_at,
            pdf_mega_path=t.pdf_mega_path,
        )
        for t in talents
    ]


# ─── Edit + history (TKT-0167) ───────────────────────────────────────────────
# Lets the team correct a paperwork field after the fact (e.g. fix a typo in
# an address) without re-signing. Every edit is captured as a row in
# compliance_signatures_history via the AFTER UPDATE trigger, so paperwork
# can be viewed "as of date X".


# Editable fields. Identifiers (id, shoot_id, talent_role, talent_slug,
# scene_id, studio, signed_at, contract_version, signature_image_path,
# pdf_*) are intentionally NOT editable — they're the audit anchor.
_EDITABLE_FIELDS: set[str] = {
    "talent_display",
    "legal_name", "business_name",
    "tax_classification", "llc_class", "other_classification",
    "exempt_payee_code", "fatca_code",
    "tin_type", "tin",
    "dob", "place_of_birth",
    "street_address", "city_state_zip", "phone", "email",
    "id1_type", "id1_number", "id2_type", "id2_number",
    "stage_names", "professional_names",
    "nicknames_aliases", "previous_legal_names",
}


class SignatureEditRequest(BaseModel):
    """Partial update — only fields included in `changes` are written.

    Fields not in `_EDITABLE_FIELDS` are ignored server-side. Pass an
    optional `reason` to surface in the history audit trail.
    """
    changes: dict[str, str]
    reason: str = ""


class SignatureRow(BaseModel):
    id: int
    shoot_id: str
    shoot_date: str
    scene_id: str = ""
    studio: str = ""
    talent_role: str
    talent_slug: str
    talent_display: str
    legal_name: str
    business_name: str = ""
    tax_classification: str
    llc_class: str = ""
    other_classification: str = ""
    exempt_payee_code: str = ""
    fatca_code: str = ""
    tin_type: str
    tin: str
    dob: str
    place_of_birth: str
    street_address: str
    city_state_zip: str
    phone: str
    email: str
    id1_type: str
    id1_number: str
    id2_type: str = ""
    id2_number: str = ""
    stage_names: str = ""
    professional_names: str = ""
    nicknames_aliases: str = ""
    previous_legal_names: str = ""
    signed_at: str
    contract_version: str
    pdf_mega_path: str = ""
    created_at: str = ""


class SignatureHistoryEntry(SignatureRow):
    history_id: int
    snapshot_at: str
    edited_by: str = ""
    edit_reason: str = ""


@router.get("/signatures/{signature_id}", response_model=SignatureRow)
async def get_signature(signature_id: int, user: CurrentUser):
    """Return one signature row for editing."""
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM compliance_signatures WHERE id=?",
            (signature_id,),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="signature not found")
    return SignatureRow(**dict(row))


@router.patch("/signatures/{signature_id}", response_model=SignatureRow)
async def edit_signature(
    signature_id: int,
    body: SignatureEditRequest,
    user: CurrentUser,
):
    """Patch one or more editable fields. Trigger writes the prior state
    to `compliance_signatures_history`. Identifiers are protected — any
    non-editable key is silently dropped."""
    safe = {k: v for k, v in body.changes.items() if k in _EDITABLE_FIELDS}
    if not safe:
        raise HTTPException(status_code=400, detail="no editable fields in changes")
    set_clause = ", ".join(f"{k}=?" for k in safe)
    params = [*safe.values(), signature_id]

    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM compliance_signatures WHERE id=?",
            (signature_id,),
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="signature not found")
        # Capture the editor + reason for the history row by stuffing them
        # into a temp table the trigger doesn't touch — but SQLite has no
        # session vars, so we just write a follow-up UPDATE on the most
        # recent history row right after the edit. Race-free because we
        # serialize within this connection.
        conn.execute(
            f"UPDATE compliance_signatures SET {set_clause} WHERE id=?",
            params,
        )
        editor = (user or {}).get("email", "") if isinstance(user, dict) else ""
        if editor or body.reason:
            conn.execute(
                "UPDATE compliance_signatures_history "
                "SET edited_by=?, edit_reason=? "
                "WHERE history_id=(SELECT MAX(history_id) FROM compliance_signatures_history "
                "                  WHERE signature_id=?)",
                (editor, body.reason, signature_id),
            )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM compliance_signatures WHERE id=?",
            (signature_id,),
        ).fetchone()
    return SignatureRow(**dict(row))


@router.get("/signatures/{signature_id}/history",
            response_model=list[SignatureHistoryEntry])
async def get_signature_history(signature_id: int, user: CurrentUser):
    """Return every prior state of this signature, newest first.
    Each row reflects the state BEFORE the edit at `snapshot_at`."""
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM compliance_signatures_history "
            "WHERE signature_id=? ORDER BY history_id DESC",
            (signature_id,),
        ).fetchall()
    return [SignatureHistoryEntry(**dict(r)) for r in rows]


@router.get("/signatures/{signature_id}/as-of",
            response_model=SignatureRow)
async def get_signature_as_of(
    signature_id: int,
    at: str,            # ISO timestamp — return state at this moment
    user: CurrentUser,
):
    """Return the signature as it existed at a specific moment.

    Strategy:
      * If a history row was snapshotted AFTER `at`, the OLDEST such row's
        BEFORE-state IS the answer (it captures the state at that moment).
      * Otherwise the current row is the answer (no edits since `at`).

    Lets the audit / paperwork-correction UI show "this is what we had
    on file when X happened."
    """
    with get_db() as conn:
        # find the oldest history row with snapshot_at >= at
        hist = conn.execute(
            "SELECT * FROM compliance_signatures_history "
            "WHERE signature_id=? AND snapshot_at >= ? "
            "ORDER BY snapshot_at ASC LIMIT 1",
            (signature_id, at),
        ).fetchone()
        if hist:
            d = dict(hist)
            # SignatureRow expects id, not signature_id
            d["id"] = d.pop("signature_id")
            d.pop("history_id", None)
            d.pop("snapshot_at", None)
            d.pop("edited_by", None)
            d.pop("edit_reason", None)
            return SignatureRow(**d)
        row = conn.execute(
            "SELECT * FROM compliance_signatures WHERE id=?",
            (signature_id,),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="signature not found")
    return SignatureRow(**dict(row))


@router.get("/signatures/{signature_id}/render-pdf")
async def render_signature_pdf(
    signature_id: int,
    user: CurrentUser,
    date: Optional[str] = Query(None,
        description="Override the date stamped on the rendered PDF "
                    "(format: 'Apr 30, 2026'). Defaults to the row's shoot_date."),
    as_of: Optional[str] = Query(None,
        description="ISO timestamp — render the row's state as of this moment."),
):
    """Render a fresh PDF from the row's current (or as-of-date) state.

    Backbone of the "go back to a specific date" feature: combine
    `as_of` (which row state to render) with `date` (which date to
    stamp on it) to reproduce any historical artifact.

    Streams the PDF inline so the browser can preview or save.
    """
    from fastapi.responses import StreamingResponse

    if as_of:
        # Reuse the same logic as get_signature_as_of — earliest history
        # row >= as_of, else current.
        with get_db() as conn:
            hist = conn.execute(
                "SELECT * FROM compliance_signatures_history "
                "WHERE signature_id=? AND snapshot_at >= ? "
                "ORDER BY snapshot_at ASC LIMIT 1",
                (signature_id, as_of),
            ).fetchone()
            if hist:
                d = dict(hist)
                row = d
            else:
                cur = conn.execute(
                    "SELECT * FROM compliance_signatures WHERE id=?",
                    (signature_id,),
                ).fetchone()
                if not cur:
                    raise HTTPException(status_code=404, detail="signature not found")
                row = dict(cur)
    else:
        with get_db() as conn:
            cur = conn.execute(
                "SELECT * FROM compliance_signatures WHERE id=?",
                (signature_id,),
            ).fetchone()
        if not cur:
            raise HTTPException(status_code=404, detail="signature not found")
        row = dict(cur)

    pdf_bytes = _render_pdf_from_signature(row, date_override=date)
    slug = row.get("talent_slug") or "talent"
    shoot_date = (row.get("shoot_date") or "")[:10].replace("-", "")
    filename = f"{slug}-{shoot_date or 'rendered'}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


# ─── Pre-fill for returning female talent (TKT-0167) ────────────────────────
# Different from male auto-sign — females review + sign each visit, but if
# the same female has signed within the last 12 months we surface her prior
# answers as form defaults so she only has to confirm/correct, not retype.


class TalentPrefill(BaseModel):
    """Returned by /talent/{slug}/recent-prefill. None of these are PII the
    UI hadn't already seen — they're echoed back for a returning talent."""
    found: bool
    source_shoot_id: str = ""
    source_signed_at: str = ""
    legal_name: str = ""
    business_name: str = ""
    tax_classification: str = ""
    llc_class: str = ""
    other_classification: str = ""
    exempt_payee_code: str = ""
    fatca_code: str = ""
    tin_type: str = ""
    tin: str = ""
    dob: str = ""
    place_of_birth: str = ""
    street_address: str = ""
    city_state_zip: str = ""
    phone: str = ""
    email: str = ""
    id1_type: str = ""
    id1_number: str = ""
    id2_type: str = ""
    id2_number: str = ""
    stage_names: str = ""


@router.get("/talent/{talent_slug}/recent-prefill", response_model=TalentPrefill)
async def talent_recent_prefill(
    talent_slug: str,
    user: CurrentUser,
    role: str = "female",
    within_days: int = 365,
):
    """Return the most recent compliance_signatures row for a talent, if it
    falls inside the lookback window. Used by the UI to pre-populate the
    form on return shoots — the talent still reviews + signs.

    ``within_days`` defaults to 365 (12 months). The talent can update any
    field before signing; their answers go into a fresh row, leaving the
    prior one untouched.
    """
    cutoff = (datetime.now(timezone.utc) - timedelta(days=within_days)).date().isoformat()
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM compliance_signatures "
            "WHERE talent_role=? AND talent_slug=? "
            "  AND legal_name != '' AND legal_name IS NOT NULL "
            "  AND signed_at >= ? "
            "ORDER BY signed_at DESC LIMIT 1",
            (role, talent_slug, cutoff),
        ).fetchone()
    if not row:
        return TalentPrefill(found=False)
    d = dict(row)
    return TalentPrefill(
        found=True,
        source_shoot_id=d.get("shoot_id", ""),
        source_signed_at=d.get("signed_at", ""),
        legal_name=d.get("legal_name", ""),
        business_name=d.get("business_name", ""),
        tax_classification=d.get("tax_classification", ""),
        llc_class=d.get("llc_class", ""),
        other_classification=d.get("other_classification", ""),
        exempt_payee_code=d.get("exempt_payee_code", ""),
        fatca_code=d.get("fatca_code", ""),
        tin_type=d.get("tin_type", ""),
        tin=d.get("tin", ""),
        dob=d.get("dob", ""),
        place_of_birth=d.get("place_of_birth", ""),
        street_address=d.get("street_address", ""),
        city_state_zip=d.get("city_state_zip", ""),
        phone=d.get("phone", ""),
        email=d.get("email", ""),
        id1_type=d.get("id1_type", ""),
        id1_number=d.get("id1_number", ""),
        id2_type=d.get("id2_type", ""),
        id2_number=d.get("id2_number", ""),
        stage_names=d.get("stage_names", ""),
    )


# ─── Auto-sign for males whose paperwork is already on file (TKT-0167) ──────
# Males in MALE_TPLS shoot back-to-back with the same pre-filled paperwork
# (W-9 + 2257 stay constant per talent for years). The team shouldn't have
# to re-collect a signature each time. This endpoint clones the male's most
# recent compliance_signatures row into the target shoot, so the UI sees
# them as "signed" without an iPad round trip.


class AutoSignTalentResult(BaseModel):
    shoot_id: str
    talent_role: str
    talent_slug: str
    source_shoot_id: str = ""    # the prior record we cloned, if any
    created_signature_id: int = 0
    ids_copied: list[str] = []   # filenames copied from the prior shoot folder
    skipped_reason: str = ""     # set if no prior record exists


def _copy_prior_ids_to_shoot_folder(
    prior_shoot_date_iso: str,
    prior_female: str,
    prior_male: str,
    talent_role: str,
    talent_slug: str,
    dest_folder_id: str,
    existing_dest_lower: set[str],
    token: str,
) -> list[str]:
    """Find the prior shoot's Drive folder, copy that talent's IDs into
    the destination folder. Idempotent — skips already-present filenames.
    Returns list of newly-copied destination filenames.
    """
    from datetime import datetime as _dt
    try:
        prior_date = _dt.fromisoformat(prior_shoot_date_iso[:10]).date()
    except Exception:
        return []
    info = _get_shoot_folder(prior_date, prior_female, prior_male or "", token)
    if not info:
        return []
    prior_folder_id, _ = info
    prior_files = _list_folder_files(prior_folder_id, token)

    slug_lower = talent_slug.lower()
    copied: list[str] = []
    for f in prior_files:
        name = f.get("name", "") or ""
        if not name:
            continue
        # Match files belonging to this talent — slug-prefixed filenames
        # like "LeanaLovings-id-front.jpg" or "DannySteele-bunny-ear.HEIC".
        name_collapsed = name.replace(" ", "").replace("-", "").lower()
        if slug_lower not in name_collapsed:
            continue
        # Match ID-like filenames (front/back/bunny/id) — skip random photos
        nm_lower = name.lower()
        if not any(k in nm_lower for k in ("front", "back", "bunny", "-id-", "id_")):
            continue
        if name.lower() in existing_dest_lower:
            continue
        try:
            _copy_file(f["id"], dest_folder_id, name, token)
            copied.append(name)
            existing_dest_lower.add(name.lower())
        except Exception as exc:
            _log.warning("auto-sign ID copy %s failed: %s", name, exc)
    return copied


def _do_auto_sign(shoot_id: str, talent_role: str) -> AutoSignTalentResult:
    """Shared implementation for both female and male auto-sign.

    Clones the talent's most recent populated compliance_signatures row
    into this shoot, then copies their ID photos from that prior shoot's
    Drive folder into the current shoot's Drive folder.

    Idempotent — UNIQUE(shoot_id, talent_role, talent_slug) means re-calling
    overwrites the existing row (history trigger captures the prior state).
    """
    if talent_role not in ("female", "male"):
        raise HTTPException(status_code=400, detail="role must be 'female' or 'male'")

    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="shoot not found")

    talent_display = shoot.female_talent if talent_role == "female" else (shoot.male_talent or "")
    if not talent_display:
        raise HTTPException(status_code=400, detail=f"shoot has no {talent_role} talent")
    talent_slug = talent_display.replace(" ", "")

    bg_scenes = [sc for sc in shoot.scenes if sc.scene_type.upper() in COMPLIANCE_SCENE_TYPES]
    primary = bg_scenes[0] if bg_scenes else None
    scene_id = primary.scene_id if primary else ""
    studio = primary.studio if primary else ""

    # Find the talent's most recent prior populated row.
    with get_db() as conn:
        prior = conn.execute(
            "SELECT * FROM compliance_signatures "
            "WHERE talent_role=? AND talent_slug=? "
            "  AND legal_name != '' AND legal_name IS NOT NULL "
            "  AND shoot_id != ? "
            "ORDER BY signed_at DESC LIMIT 1",
            (talent_role, talent_slug, shoot_id),
        ).fetchone()

    if not prior:
        return AutoSignTalentResult(
            shoot_id=shoot_id,
            talent_role=talent_role,
            talent_slug=talent_slug,
            skipped_reason=f"No prior signature on file for {talent_slug}. "
                           "Run /admin/bulk-import-from-drive or have the "
                           "talent sign once in-person.",
        )

    sig_id = upsert_signature(
        shoot_id=shoot_id,
        shoot_date=shoot.shoot_date,
        scene_id=scene_id,
        studio=studio,
        talent_role=talent_role,
        talent_slug=talent_slug,
        talent_display=talent_display,
        legal_name=prior["legal_name"],
        business_name=prior["business_name"] or "",
        tax_classification=prior["tax_classification"] or "individual",
        llc_class=prior["llc_class"] or "",
        other_classification=prior["other_classification"] or "",
        exempt_payee_code=prior["exempt_payee_code"] or "",
        fatca_code=prior["fatca_code"] or "",
        tin_type=prior["tin_type"] or "ssn",
        tin=prior["tin"] or "",
        dob=prior["dob"] or "",
        place_of_birth=prior["place_of_birth"] or "",
        street_address=prior["street_address"] or "",
        city_state_zip=prior["city_state_zip"] or "",
        phone=prior["phone"] or "",
        email=prior["email"] or "",
        id1_type=prior["id1_type"] or "",
        id1_number=prior["id1_number"] or "",
        id2_type=prior["id2_type"] or "",
        id2_number=prior["id2_number"] or "",
        stage_names=prior["stage_names"] or talent_display,
        professional_names=prior["professional_names"] or "",
        nicknames_aliases=prior["nicknames_aliases"] or "",
        previous_legal_names=prior["previous_legal_names"] or "",
        signature_image_path=prior["signature_image_path"] or "auto-fill-from-prior",
        signed_ip="",
        signed_user_agent=f"auto-sign-{talent_role}",
        signed_by_user=f"auto-fill:from-{prior['shoot_id']}",
        pdf_local_path="",
        pdf_mega_path="",
    )

    # Auto-copy IDs from the prior shoot's Drive folder. Best-effort —
    # failures don't roll back the signature.
    ids_copied: list[str] = []
    try:
        token = _get_drive_rw_token()
        if token:
            shoot_date_obj = _parse_shoot_date(shoot.shoot_date)
            if shoot_date_obj:
                # Resolve current shoot folder + its existing files.
                month_id = _find_or_create_folder(LEGAL_ROOT_FOLDER,
                                                  shoot_date_obj.strftime("%B"), token)
                date_code = shoot_date_obj.strftime("%m%d%y")
                female_slug = shoot.female_talent.replace(" ", "")
                male_slug = (shoot.male_talent or "").replace(" ", "")
                folder_name = f"{date_code}-{female_slug}" + (f"-{male_slug}" if male_slug else "")
                folder_id = _find_or_create_folder(month_id, folder_name, token)
                existing_lower = {f.get("name", "").lower()
                                  for f in _list_folder_files(folder_id, token)}

                # Look up the prior shoot's female/male talent so we can find
                # its folder. The shoot_id stores it but the field naming
                # convention isn't 1:1, so look up by date in the shoots window.
                prior_shoot = next(
                    (s for s in shoots if s.shoot_id == prior["shoot_id"]),
                    None,
                )
                if prior_shoot:
                    ids_copied = _copy_prior_ids_to_shoot_folder(
                        prior_shoot.shoot_date,
                        prior_shoot.female_talent,
                        prior_shoot.male_talent or "",
                        talent_role, talent_slug,
                        folder_id, existing_lower, token,
                    )
    except Exception as exc:
        _log.warning("auto-sign ID copy chain failed: %s", exc)

    return AutoSignTalentResult(
        shoot_id=shoot_id,
        talent_role=talent_role,
        talent_slug=talent_slug,
        source_shoot_id=prior["shoot_id"],
        created_signature_id=sig_id,
        ids_copied=ids_copied,
    )


@router.post("/shoots/{shoot_id}/auto-sign-male", response_model=AutoSignTalentResult)
async def auto_sign_male(shoot_id: str, user: CurrentUser):
    """Clone the male's most recent compliance_signatures row + IDs into
    this shoot. See `_do_auto_sign` for full semantics."""
    return _do_auto_sign(shoot_id, "male")


@router.post("/shoots/{shoot_id}/auto-sign-female", response_model=AutoSignTalentResult)
async def auto_sign_female(shoot_id: str, user: CurrentUser):
    """Clone the female's most recent compliance_signatures row + IDs into
    this shoot. Use when a returning female has shot back-to-back and
    nothing on her paperwork has changed.

    Same flow as auto-sign-male but for the female role. The talent-prefill
    UI (which keeps the sign step) is the right choice when the talent
    might want to update fields; this is the "she just shot yesterday,
    same paperwork" fast path.
    """
    return _do_auto_sign(shoot_id, "female")


# ─── Legacy Drive paperwork import (TKT-0152) ────────────────────────────────
# Wires already-signed Drive PDFs into compliance_signatures so a shoot whose
# talent finished paperwork in the legacy Drive flow shows complete in the Hub
# without forcing them to re-sign on the iPad. The original Drive PDF — with
# its real drawn signatures — is the legal artifact: we copy it byte-for-byte
# to MEGA Grail/{Studio}/{scene_id}/Legal/ and link to it from the Hub. We do
# NOT regenerate the agreement, do NOT fabricate a signature, and do NOT pull
# PII into the Hub UI. The compliance_signatures row is a thin index pointing
# at the Drive original; W-9/2257 fields stay empty and are sourced from the
# linked PDF when needed.


_DRIVE_FOLDER_ID_RE = re.compile(r"/folders/([A-Za-z0-9_\-]+)")


def _parse_drive_folder_id(url_or_id: str) -> str:
    """Accept a Drive folder URL or a bare folder id."""
    s = (url_or_id or "").strip()
    if not s:
        raise HTTPException(status_code=400, detail="folder_url required")
    m = _DRIVE_FOLDER_ID_RE.search(s)
    if m:
        return m.group(1)
    # Looks like a bare ID already
    if re.fullmatch(r"[A-Za-z0-9_\-]{20,}", s):
        return s
    raise HTTPException(status_code=400, detail=f"Cannot parse Drive folder from {s!r}")


def _drive_download_bytes(file_id: str, token: str) -> bytes:
    req = urllib.request.Request(
        f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media",
        headers={"Authorization": f"Bearer {token}"},
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read()


def _import_placeholder_signature_png() -> bytes:
    """Tiny labeled PNG used as the signature_image_path for imports.

    The DB schema requires a non-null path, but we don't want to fabricate a
    cursive-looking signature. This image plainly says the signature lives
    in the linked PDF, so anyone inspecting the audit folder is immediately
    pointed at the real artifact.
    """
    try:
        from PIL import Image, ImageDraw, ImageFont  # type: ignore
    except Exception:
        # Fallback: smallest valid PNG (1×1 transparent)
        import base64
        return base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNkYAAAAAYAAjCB0C8AAAAASUVORK5CYII="
        )
    img = Image.new("RGB", (520, 110), "white")
    d = ImageDraw.Draw(img)
    try:
        font_path = next(
            (p for p in [
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "C:/Windows/Fonts/segoeui.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "/Library/Fonts/Arial.ttf",
            ] if Path(p).exists()),
            None,
        )
        head = ImageFont.truetype(font_path, 18) if font_path else ImageFont.load_default()
        body = ImageFont.truetype(font_path, 13) if font_path else ImageFont.load_default()
    except Exception:
        head = body = ImageFont.load_default()
    d.text((22, 22), "Signature on file in linked PDF", fill=(40, 40, 40), font=head)
    d.text((22, 56), "Imported from Drive — see attached agreement.", fill=(120, 120, 120), font=body)
    d.text((22, 78), "(no in-Hub signature was captured for this record)", fill=(160, 160, 160), font=body)
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


class DriveImportRequest(BaseModel):
    """Pull every PDF in a Drive folder into the Hub for the shoot it
    targets. Filenames matching the talent slug prefix are matched to roles."""
    folder_url: str
    imported_from_date: str = ""   # original sign date for audit trail


class DriveImportTalentResult(BaseModel):
    talent_role: str
    talent_slug: str
    pdf_local_path: str
    pdf_mega_path: str
    bytes_copied: int


class DriveImportResult(BaseModel):
    shoot_id: str
    imported: list[DriveImportTalentResult] = []
    skipped: list[str] = []
    errors: list[str] = []


@router.post("/shoots/{shoot_id}/import-from-drive", response_model=DriveImportResult)
async def import_from_drive(
    shoot_id: str,
    user: CurrentUser,
    request: Request,
    body: DriveImportRequest,
):
    """Import already-signed Drive PDFs into compliance_signatures.

    Server-side flow (no PII passes through the Hub UI):
      1. List PDFs in the Drive folder
      2. Match each to female/male by filename prefix
      3. Copy bytes to compliance_pdfs/{shoot_date}/{Slug}-{date}.pdf
      4. Push the same bytes to MEGA Grail/{Studio}/{scene_id}/Legal/
      5. Insert a compliance_signatures row pointing at the saved PDF; W-9
         and 2257 fields are left empty (the linked PDF is the legal record).
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    bg_scenes = [sc for sc in shoot.scenes if sc.scene_type.upper() in COMPLIANCE_SCENE_TYPES]
    if not bg_scenes:
        raise HTTPException(status_code=400, detail="Shoot has no BG/BGCP scene")
    primary = bg_scenes[0]
    scene_id = primary.scene_id or ""
    studio = primary.studio or ""

    shoot_date_obj = _parse_shoot_date(shoot.shoot_date)
    if not shoot_date_obj:
        raise HTTPException(status_code=400, detail="Invalid shoot date")

    folder_id = _parse_drive_folder_id(body.folder_url)
    token = _get_drive_token()
    if not token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")

    files = await asyncio.to_thread(_list_folder_files, folder_id, token)
    pdfs = [f for f in files if (f.get("name") or "").lower().endswith(".pdf")]
    if not pdfs:
        raise HTTPException(status_code=404, detail="No PDFs found in folder")

    # Match each PDF to either talent by slug prefix in filename
    female_slug = shoot.female_talent.replace(" ", "")
    male_slug = (shoot.male_talent or "").replace(" ", "")
    matches: dict[str, dict] = {}
    for f in pdfs:
        name = (f.get("name") or "").replace(" ", "")
        low = name.lower()
        if female_slug and low.startswith(female_slug.lower()):
            matches.setdefault("female", f)
        elif male_slug and low.startswith(male_slug.lower()):
            matches.setdefault("male", f)

    result = DriveImportResult(shoot_id=shoot.shoot_id)
    if not matches:
        result.errors.append(
            f"No PDFs matched talent slugs ({female_slug!r}, {male_slug or '—'!r})"
        )
        return result

    audit_user = (
        f"legacy_import:drive"
        + (f":{body.imported_from_date}" if body.imported_from_date else "")
    )
    placeholder_png = _import_placeholder_signature_png()
    signed_at_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    date_code = shoot_date_obj.strftime("%m%d%y")

    for role, drive_file in matches.items():
        slug = female_slug if role == "female" else male_slug
        display = shoot.female_talent if role == "female" else shoot.male_talent
        try:
            pdf_bytes = await asyncio.to_thread(_drive_download_bytes, drive_file["id"], token)
        except Exception as exc:
            result.errors.append(f"{slug}: download failed — {exc}")
            continue

        # Save locally
        pdf_name = f"{slug}-{date_code}.pdf"
        pdf_path = _legal_pdf_dir() / shoot.shoot_date / pdf_name
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        pdf_path.write_bytes(pdf_bytes)

        # Save the signature placeholder
        sig_path = _signature_dir() / f"{shoot.shoot_date}-{slug}-{role}.png"
        sig_path.write_bytes(placeholder_png)

        # Push to MEGA S4
        mega_remote = ""
        mega_err: Optional[str] = None
        if scene_id:
            s4_key = s4_client.key_for(scene_id, "Legal", pdf_name)
            bucket = s4_client.STUDIO_BUCKETS[s4_client._STUDIO_ALIASES.get(studio, studio).upper()]
            mega_remote = f"s3://{bucket}/{s4_key}"
            mega_err = await asyncio.to_thread(_push_to_mega, pdf_path, studio, s4_key)
            if mega_err:
                _log.warning("S4 push failed for %s: %s", pdf_name, mega_err)

        # Insert compliance_signatures row — W-9/2257 fields stay empty, the
        # linked PDF is the legal record. legal_name defaults to the talent
        # display so the picker shows something readable.
        upsert_signature(
            shoot_id=shoot.shoot_id,
            shoot_date=shoot.shoot_date,
            scene_id=scene_id,
            studio=studio,
            talent_role=role,
            talent_slug=slug,
            talent_display=display,
            legal_name=display,           # placeholder; real name lives in PDF
            business_name="",
            tax_classification="individual",
            llc_class="",
            other_classification="",
            exempt_payee_code="",
            fatca_code="",
            tin_type="ssn",
            tin="",
            dob="",
            place_of_birth="",
            street_address="",
            city_state_zip="",
            phone="",
            email="",
            id1_type="",
            id1_number="",
            id2_type="",
            id2_number="",
            stage_names=display,
            professional_names="",
            nicknames_aliases="",
            previous_legal_names="",
            signature_image_path=str(sig_path.relative_to(get_settings().base_dir)),
            signed_ip=request.client.host if request.client else "",
            signed_user_agent="legacy-import",
            signed_by_user=audit_user,
            pdf_local_path=str(pdf_path),
            pdf_mega_path=mega_remote if (mega_remote and not mega_err) else "",
        )

        result.imported.append(DriveImportTalentResult(
            talent_role=role,
            talent_slug=slug,
            pdf_local_path=str(pdf_path),
            pdf_mega_path=mega_remote if (mega_remote and not mega_err) else "",
            bytes_copied=len(pdf_bytes),
        ))

    if female_slug and "female" not in matches:
        result.skipped.append(f"female ({female_slug})")
    if male_slug and "male" not in matches:
        result.skipped.append(f"male ({male_slug})")
    return result


# ─── Server-persisted photos (TKT-0151) ──────────────────────────────────────
# Photos no longer require the Drive folder, talent signatures, or any other
# precondition. They are saved to the local filesystem under
# compliance_photos/{shoot_id}/, indexed in the compliance_photos table, and
# (when scene_id is known) pushed to MEGA Grail/{Studio}/{scene_id}/Legal/.
# The same shoot can have photos uploaded across multiple visits — each slot
# is keyed by (shoot_id, slot_id) so re-uploading replaces the prior file.


class PhotoSummary(BaseModel):
    slot_id: str
    talent_role: str
    label: str
    mime_type: str
    file_size: int
    uploaded_at: str
    mega_path: str
    url: str           # GET URL for the photo bytes (used as <img src>)


def _photo_dir(shoot_id: str) -> Path:
    """Local on-disk storage for a shoot's compliance photos."""
    safe = re.sub(r"[^A-Za-z0-9._\-]", "_", shoot_id)
    p = Path(get_settings().base_dir) / "compliance_photos" / safe
    p.mkdir(parents=True, exist_ok=True)
    return p


def _photo_url(shoot_id: str, slot_id: str) -> str:
    return (
        "/api/compliance/shoots/"
        + urllib.parse.quote(shoot_id, safe="")
        + "/photos-v2/"
        + urllib.parse.quote(slot_id, safe="")
    )


def _ensure_extension(label: str, mime_type: str) -> str:
    """Make sure label has a sensible extension — Hub sends labels with one
    already, but we don't trust the client."""
    if any(label.lower().endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".mp4", ".mov", ".webm")):
        return label
    if mime_type.startswith("video/"):
        return label + ".mp4"
    return label + ".jpg"


@router.get("/shoots/{shoot_id}/photos-v2", response_model=list[PhotoSummary])
async def list_compliance_photos(shoot_id: str, user: CurrentUser):
    """Return every photo persisted server-side for a shoot."""
    return [
        PhotoSummary(
            slot_id=p.slot_id,
            talent_role=p.talent_role,
            label=p.label,
            mime_type=p.mime_type,
            file_size=p.file_size,
            uploaded_at=p.uploaded_at,
            mega_path=p.mega_path,
            url=_photo_url(shoot_id, p.slot_id),
        )
        for p in _db_list_photos(shoot_id)
    ]


@router.get("/shoots/{shoot_id}/photos-v2/{slot_id}")
async def get_compliance_photo(shoot_id: str, slot_id: str, user: CurrentUser):
    """Serve the bytes of a persisted photo (used as <img src>)."""
    from fastapi.responses import Response as FastAPIResponse

    p = _db_get_photo(shoot_id, slot_id)
    if not p:
        raise HTTPException(status_code=404, detail="Photo not found")
    fp = Path(p.local_path)
    if not fp.exists():
        raise HTTPException(status_code=404, detail="Photo file missing on disk")
    return FastAPIResponse(
        content=fp.read_bytes(),
        media_type=p.mime_type,
        headers={
            "Content-Disposition": f"inline; filename={p.label}",
            "Cache-Control": "private, max-age=300",
        },
    )


@router.post("/shoots/{shoot_id}/photos-v2", response_model=PhotoSummary)
async def upload_compliance_photo(
    shoot_id: str,
    user: CurrentUser,
    slot_id: str = Form(...),
    label: str = Form(...),
    talent_role: str = Form(default=""),
    file: UploadFile = File(...),
):
    """Persist a single photo for a shoot.

    Saves to local disk + DB index + MEGA (when scene_id is known). Replaces
    any prior upload for the same (shoot_id, slot_id). Photos work even when
    talent has not signed yet — that's the whole point of this endpoint.
    """
    from_d, to_d = _window()
    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    shoot = next((s for s in shoots if s.shoot_id == shoot_id), None)
    if not shoot:
        raise HTTPException(status_code=404, detail="Shoot not found")

    bg_scenes = [sc for sc in shoot.scenes if sc.scene_type.upper() in COMPLIANCE_SCENE_TYPES]
    primary = bg_scenes[0] if bg_scenes else None
    scene_id = (primary.scene_id if primary else "") or ""
    studio   = (primary.studio   if primary else "") or ""

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty upload")

    mime = file.content_type or ("video/mp4" if label.lower().endswith((".mp4", ".mov", ".webm")) else "image/jpeg")
    safe_label = _ensure_extension(label, mime)
    safe_slot = re.sub(r"[^A-Za-z0-9._\-]", "_", slot_id) or "slot"

    # Replace any existing photo for this (shoot_id, slot_id) on disk before
    # writing the new bytes — prevents stale extensions if mime changed.
    existing = _db_get_photo(shoot_id, slot_id)
    if existing:
        try:
            Path(existing.local_path).unlink(missing_ok=True)
        except Exception as exc:
            _log.debug("photo unlink old failed: %s", exc)

    # Write the new file
    dest = _photo_dir(shoot_id) / f"{safe_slot}__{safe_label}"
    dest.write_bytes(content)

    # Push to MEGA S4 when we know the scene
    mega_path = ""
    if scene_id and studio:
        s4_key = s4_client.key_for(scene_id, "Legal", safe_label)
        bucket = s4_client.STUDIO_BUCKETS[s4_client._STUDIO_ALIASES.get(studio, studio).upper()]
        err = await asyncio.to_thread(_push_to_mega, dest, studio, s4_key)
        if err:
            _log.warning("S4 push failed for photo %s: %s", safe_label, err)
        else:
            mega_path = f"s3://{bucket}/{s4_key}"

    _db_upsert_photo(
        shoot_id=shoot_id,
        shoot_date=shoot.shoot_date,
        scene_id=scene_id,
        studio=studio,
        slot_id=slot_id,
        talent_role=(talent_role or "").lower(),
        label=safe_label,
        mime_type=mime,
        file_size=len(content),
        local_path=str(dest),
        mega_path=mega_path,
        uploaded_by=getattr(user, "email", "") or "",
    )

    return PhotoSummary(
        slot_id=slot_id,
        talent_role=(talent_role or "").lower(),
        label=safe_label,
        mime_type=mime,
        file_size=len(content),
        uploaded_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        mega_path=mega_path,
        url=_photo_url(shoot_id, slot_id),
    )


@router.delete("/shoots/{shoot_id}/photos-v2/{slot_id}")
async def delete_compliance_photo(shoot_id: str, slot_id: str, user: CurrentUser):
    """Drop a photo from disk + DB. MEGA copy is left in place — the user
    can resync if they want it gone there."""
    p = _db_delete_photo(shoot_id, slot_id)
    if p:
        try:
            Path(p.local_path).unlink(missing_ok=True)
        except Exception as exc:
            _log.debug("photo unlink failed: %s", exc)
    return {"ok": True}


# ─── Admin W-9 export (TKT-0153) ─────────────────────────────────────────────
# Admin-only spreadsheet export of every compliance_signatures row, formatted
# for handoff to the accountant. The export is generated on the server and
# streamed as an .xlsx — admins can filter by date range and studio. SSN/EIN
# are formatted with dashes for readability.


def _format_tin(tin_type: str, tin: str) -> str:
    digits = "".join(c for c in (tin or "") if c.isdigit())
    if not digits:
        return ""
    if tin_type == "ein" and len(digits) == 9:
        return f"{digits[:2]}-{digits[2:]}"
    if tin_type == "ssn" and len(digits) == 9:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return digits


def _format_phone(phone: str) -> str:
    digits = "".join(c for c in (phone or "") if c.isdigit())
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"+1 ({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return phone or ""


def _tax_class_label(c: str, llc: str = "", other: str = "") -> str:
    labels = {
        "individual":   "Individual / Sole Proprietor",
        "c_corp":       "C Corporation",
        "s_corp":       "S Corporation",
        "partnership":  "Partnership",
        "trust_estate": "Trust / Estate",
    }
    if c in labels:
        return labels[c]
    if c == "llc":
        return f"LLC ({llc or '?'})"
    if c == "other":
        return f"Other — {other or '?'}"
    return c or ""


def _build_w9_xlsx(records: list[W9Record], date_from: str, date_to: str, studio: str) -> bytes:
    """Render compliance_signatures rows into a polished Excel workbook.

    Single sheet with frozen header, banded rows, and column widths sized to
    fit content. SSN/EIN/phone are pre-formatted; the original signed PDF
    path is included so the accountant can pull source artifacts on demand."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "W-9 Records"

    # Header palette — neutral graphite + Eclatech lime accent
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    title_font  = Font(name="Calibri", size=16, bold=True, color="1A1A1A")
    subtitle_font = Font(name="Calibri", size=10, color="666666", italic=True)
    body_font   = Font(name="Calibri", size=10, color="222222")
    mono_font   = Font(name="Consolas",  size=10, color="222222")
    band_fill   = PatternFill("solid", fgColor="F6F6F4")
    thin_border = Border(
        bottom=Side(style="thin", color="DCDCDC"),
        right=Side(style="thin", color="ECECEC"),
    )

    # Title block — three rows above the data
    today = datetime.now(timezone.utc).strftime("%b %-d, %Y") if hasattr(datetime, "strftime") else datetime.now(timezone.utc).strftime("%b %d, %Y")
    ws.cell(row=1, column=1, value="Eclatech LLC — W-9 / Talent Tax Records").font = title_font
    range_str = f"{date_from or '—'}  →  {date_to or '—'}"
    studio_str = studio or "All studios"
    ws.cell(row=2, column=1, value=f"{range_str}    ·    {studio_str}    ·    Generated {today} UTC    ·    {len(records)} record(s)").font = subtitle_font

    # Column definitions
    columns = [
        ("Shoot Date",       16, "shoot_date",        "date"),
        ("Studio",           12, "studio",            "text"),
        ("Scene",            10, "scene_id",          "text"),
        ("Talent (Stage)",   22, "talent_display",    "text"),
        ("Role",              8, "talent_role",       "text"),
        ("Legal Name",       28, "legal_name",        "text"),
        ("Business Name",    24, "business_name",     "text"),
        ("Tax Class",        28, "_tax_class",        "text"),
        ("TIN Type",          9, "tin_type",          "text"),
        ("TIN",              14, "_tin",              "mono"),
        ("Address",          34, "street_address",    "text"),
        ("City / State / ZIP", 24, "city_state_zip",  "text"),
        ("Phone",            16, "_phone",            "mono"),
        ("Email",            30, "email",             "text"),
        ("Signed (UTC)",     20, "signed_at",         "text"),
        ("Signed By (audit)",26, "signed_by_user",    "text"),
        ("PDF (server)",     50, "pdf_local_path",    "text"),
        ("PDF (MEGA)",       50, "pdf_mega_path",     "text"),
    ]

    HEADER_ROW = 4
    for i, (label, width, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[HEADER_ROW].height = 22
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # Body rows
    for r_idx, rec in enumerate(records, start=HEADER_ROW + 1):
        derived = {
            "_tax_class": _tax_class_label(rec.tax_classification, rec.llc_class, rec.other_classification),
            "_tin":       _format_tin(rec.tin_type, rec.tin),
            "_phone":     _format_phone(rec.phone),
        }
        banded = (r_idx - HEADER_ROW) % 2 == 0
        for c_idx, (_, _, key, kind) in enumerate(columns, start=1):
            val = derived.get(key, getattr(rec, key, ""))
            if key == "talent_role" and isinstance(val, str):
                val = val.title()
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)
            cell.font = mono_font if kind == "mono" else body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.border = thin_border
            if banded:
                cell.fill = band_fill

    # Filter handle on the header row so the accountant can sort/filter natively
    last_col = get_column_letter(len(columns))
    last_row = HEADER_ROW + max(0, len(records))
    if records:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{last_row}"

    # Footer note (one row below the table)
    if records:
        footer_row = last_row + 2
    else:
        footer_row = HEADER_ROW + 2
        ws.cell(row=HEADER_ROW + 1, column=1, value="No signed paperwork in this range.").font = subtitle_font
    ws.cell(row=footer_row, column=1, value=(
        "Source: compliance_signatures table (Hub-signed + legacy-import rows). "
        "TIN values come from the talent's W-9 entry; verify against the linked PDF before tax filing."
    )).font = subtitle_font

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.get("/admin/w9-export.xlsx")
async def export_w9_xlsx(
    _admin: dict = Depends(require_admin),
    date_from: Optional[str] = Query(default=None, alias="from"),
    date_to: Optional[str] = Query(default=None, alias="to"),
    studio: Optional[str] = Query(default=None),
):
    """Admin-only Excel download of every W-9 / talent tax record on file.

    Filters are inclusive on shoot_date. The xlsx is rendered server-side so
    PII never lives in the Hub's client memory. The header is frozen and a
    native Excel filter is applied to the data range so the accountant can
    sort / filter without external tooling."""
    from fastapi.responses import Response as FastAPIResponse

    records = list_w9_records(date_from=date_from, date_to=date_to, studio=studio)
    xlsx_bytes = _build_w9_xlsx(records, date_from or "", date_to or "", studio or "")
    range_tag = f"{date_from or 'all'}_to_{date_to or 'all'}".replace(":", "-")
    filename = f"eclatech-w9-records_{range_tag}.xlsx"
    return FastAPIResponse(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Cache-Control": "private, no-store",
        },
    )


class BulkDriveImportRequest(BaseModel):
    """Walk a Drive folder recursively and import every shoot's PDFs that
    match the conventional `MMDDYY-FemaleSlug[-MaleSlug]` folder naming.
    Useful for back-filling historical paperwork in one shot."""
    folder_url: str
    imported_from_label: str = ""   # e.g. "Drive 2026" — recorded in audit trail


class BulkDriveImportShootResult(BaseModel):
    shoot_id: str
    shoot_date: str
    folder_name: str
    talents_imported: int
    skipped_reason: str = ""


class BulkDriveImportResult(BaseModel):
    folders_seen: int
    folders_matched: int
    shoots: list[BulkDriveImportShootResult] = []
    errors: list[str] = []


_SHOOT_FOLDER_RE = re.compile(r"^(\d{2})(\d{2})(\d{2})-(.+?)(?:-(.+))?$")

# Reverse map of _map_form_to_pdf — used by the bulk importer to pull the
# actual filled values out of historical PDFs instead of saving empty
# placeholders. The existing template fills the same value into multiple
# Custom Field slots (e.g. legal_name appears in 1, 14, 16, 18, 23, 24);
# we coalesce them in `_extract_pdf_fields_for_import`.
_PDF_FIELD_TO_COL: dict[str, str] = {
    "Custom Field 1":  "legal_name",
    "Custom Field 2":  "business_name",
    "Custom Field 7":  "street_address",
    "Custom Field 9":  "city_state_zip",
    "Custom Field 17": "stage_names",
    "Custom Field 25": "dob",
    "Custom Field 26": "place_of_birth",
    "Custom Field 27": "_full_address",
    "Custom Field 28": "id1_type",
    "Custom Field 29": "id1_number",
    "Custom Field 30": "id2_type",
    "Custom Field 31": "id2_number",
    "Custom Field 32": "phone",
    "Custom Field 33": "email",
    "Custom Field 34": "stage_names",
    "Custom Field 14": "_legal_name_alt",
    "Custom Field 16": "_legal_name_alt",
    "Custom Field 18": "_legal_name_alt",
    "Custom Field 23": "_legal_name_alt",
    "Custom Field 24": "_legal_name_alt",
}


def _extract_pdf_fields_for_import(pdf_bytes: bytes) -> dict[str, str]:
    """Pull AcroForm field values out of a historical PDF.

    Returns ``{column_name: value}`` for fields recognized via
    ``_PDF_FIELD_TO_COL``. Empty dict for flattened or scanned PDFs —
    caller should fall back to the prior behavior (placeholder record).
    Errors are swallowed so a single corrupt PDF can't abort a bulk run.
    """
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    except Exception as exc:
        _log.warning("pypdf reader failed during import: %s", exc)
        return {}

    out: dict[str, str] = {}
    try:
        fields = reader.get_fields() or {}
        for name, field in fields.items():
            value = field.get("/V") if isinstance(field, dict) else None
            if value is None:
                continue
            text = str(value).strip()
            if not text:
                continue
            col = _PDF_FIELD_TO_COL.get(name)
            if col:
                out[col] = text
    except Exception as exc:
        _log.debug("get_fields() failed during import: %s", exc)

    # Annotation walk fallback — some templates expose /T directly on annots
    if not out:
        for page in reader.pages:
            if "/Annots" not in page:
                continue
            for annot in page["/Annots"]:
                try:
                    obj = annot.get_object()
                except Exception:
                    continue
                t = str(obj.get("/T", "")).strip()
                v = obj.get("/V")
                if not t or v is None:
                    continue
                text = str(v).strip()
                if not text:
                    continue
                col = _PDF_FIELD_TO_COL.get(t)
                if col and col not in out:
                    out[col] = text

    if "_legal_name_alt" in out:
        out.setdefault("legal_name", out["_legal_name_alt"])
        del out["_legal_name_alt"]
    if "_full_address" in out:
        out.setdefault("street_address", out["_full_address"])
        del out["_full_address"]
    return out


def _list_subfolders(parent_id: str, token: str) -> list[dict]:
    q = (
        f"'{parent_id}' in parents "
        "and mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    res = _drive_json(
        "https://www.googleapis.com/drive/v3/files"
        f"?q={urllib.parse.quote(q)}&fields=files(id,name)&pageSize=1000",
        token,
    )
    return (res or {}).get("files") or []


@router.post("/admin/bulk-import-from-drive", response_model=BulkDriveImportResult)
async def bulk_import_from_drive(
    body: BulkDriveImportRequest,
    request: Request,
    _admin: dict = Depends(require_admin),
):
    """Recursively walk a Drive root and import every matching shoot folder.

    The expected layout is the legacy legal-docs root: month folders →
    `MMDDYY-FemaleSlug[-MaleSlug]` shoot folders → `*Slug-MMDDYY.pdf` files.
    For each shoot folder we resolve the date + talent pair, look up the
    corresponding shoot in the local DB, and run the same per-shoot import
    used by `/import-from-drive`. Audit user is `legacy_import:bulk[:label]`.
    """
    token = _get_drive_token()
    if not token:
        raise HTTPException(status_code=503, detail="Drive credentials unavailable")
    root_id = _parse_drive_folder_id(body.folder_url)

    audit_user = (
        "legacy_import:bulk"
        + (f":{body.imported_from_label}" if body.imported_from_label else "")
    )
    placeholder_png = _import_placeholder_signature_png()
    result = BulkDriveImportResult(folders_seen=0, folders_matched=0)

    # 1. List month-level subfolders, then shoot-level subfolders inside each
    month_folders = await asyncio.to_thread(_list_subfolders, root_id, token)
    if not month_folders:
        # Fallback: maybe the user pointed at a year-level "shoots only" folder
        month_folders = [{"id": root_id, "name": ""}]

    # Cache shoots once per call. We pull a generous window so every 2026
    # shoot is reachable without hitting Drive once per query.
    today = datetime.now(timezone.utc).date()
    window_lo = today - timedelta(days=400)
    window_hi = today + timedelta(days=120)
    all_shoots = _load_shoots_window(window_lo, window_hi, include_cancelled=False)

    for month in month_folders:
        try:
            shoot_folders = await asyncio.to_thread(_list_subfolders, month["id"], token)
        except Exception as exc:
            result.errors.append(f"list {month.get('name','?')}: {exc}")
            continue

        for shoot_folder in shoot_folders:
            result.folders_seen += 1
            name = shoot_folder.get("name", "") or ""
            m = _SHOOT_FOLDER_RE.match(name.replace(" ", ""))
            if not m:
                continue
            mm, dd, yy, female_slug_raw, male_slug_raw = m.groups()
            try:
                year = 2000 + int(yy)
                shoot_date_obj = datetime(year, int(mm), int(dd)).date()
            except ValueError:
                continue
            shoot_date_iso = shoot_date_obj.isoformat()
            female_slug = (female_slug_raw or "").strip()
            male_slug = (male_slug_raw or "").strip()

            # Match the local shoot by date + female slug (slug-collapsed)
            shoot = next(
                (s for s in all_shoots
                 if s.shoot_date == shoot_date_iso
                 and s.female_talent.replace(" ", "").lower() == female_slug.lower()),
                None,
            )
            if not shoot:
                result.shoots.append(BulkDriveImportShootResult(
                    shoot_id="",
                    shoot_date=shoot_date_iso,
                    folder_name=name,
                    talents_imported=0,
                    skipped_reason="No matching shoot in local DB",
                ))
                continue
            result.folders_matched += 1

            # Pull PDFs in the shoot folder + match by talent slug prefix
            try:
                files = await asyncio.to_thread(_list_folder_files, shoot_folder["id"], token)
            except Exception as exc:
                result.errors.append(f"list {name}: {exc}")
                continue
            pdfs = [f for f in files if (f.get("name") or "").lower().endswith(".pdf")]

            bg_scenes = [sc for sc in shoot.scenes if sc.scene_type.upper() in COMPLIANCE_SCENE_TYPES]
            if not bg_scenes:
                result.shoots.append(BulkDriveImportShootResult(
                    shoot_id=shoot.shoot_id,
                    shoot_date=shoot.shoot_date,
                    folder_name=name,
                    talents_imported=0,
                    skipped_reason="Shoot has no BG scene",
                ))
                continue
            primary = bg_scenes[0]
            scene_id = primary.scene_id or ""
            studio = primary.studio or ""

            db_female = shoot.female_talent.replace(" ", "")
            db_male   = (shoot.male_talent or "").replace(" ", "")
            matches: dict[str, dict] = {}
            for f in pdfs:
                fn = (f.get("name") or "").replace(" ", "").lower()
                if db_female and fn.startswith(db_female.lower()):
                    matches.setdefault("female", f)
                elif db_male and fn.startswith(db_male.lower()):
                    matches.setdefault("male", f)

            imported = 0
            date_code = shoot_date_obj.strftime("%m%d%y")
            for role, drive_file in matches.items():
                slug = db_female if role == "female" else db_male
                display = shoot.female_talent if role == "female" else shoot.male_talent
                try:
                    pdf_bytes = await asyncio.to_thread(_drive_download_bytes, drive_file["id"], token)
                except Exception as exc:
                    result.errors.append(f"{name}/{slug}: download — {exc}")
                    continue
                pdf_name = f"{slug}-{date_code}.pdf"
                pdf_path = _legal_pdf_dir() / shoot.shoot_date / pdf_name
                pdf_path.parent.mkdir(parents=True, exist_ok=True)
                pdf_path.write_bytes(pdf_bytes)

                sig_path = _signature_dir() / f"{shoot.shoot_date}-{slug}-{role}.png"
                sig_path.write_bytes(placeholder_png)

                mega_remote = ""
                mega_err: Optional[str] = None
                if scene_id:
                    s4_key = s4_client.key_for(scene_id, "Legal", pdf_name)
                    bucket = s4_client.STUDIO_BUCKETS[s4_client._STUDIO_ALIASES.get(studio, studio).upper()]
                    mega_remote = f"s3://{bucket}/{s4_key}"
                    mega_err = await asyncio.to_thread(_push_to_mega, pdf_path, studio, s4_key)
                    if mega_err:
                        _log.warning("S4 push failed for %s: %s", pdf_name, mega_err)

                # Extract AcroForm fields from the PDF — this turns the
                # bulk import from "placeholder records" into "fully
                # populated records that can be edited after the fact"
                # (the user's actual ask). Falls back to display name +
                # empty fields for flattened/scanned PDFs that don't
                # expose AcroForm.
                extracted = _extract_pdf_fields_for_import(pdf_bytes)
                f_legal   = extracted.get("legal_name", display)
                f_biz     = extracted.get("business_name", "")
                f_dob     = extracted.get("dob", "")
                f_pob     = extracted.get("place_of_birth", "")
                f_street  = extracted.get("street_address", "")
                f_city    = extracted.get("city_state_zip", "")
                f_phone   = extracted.get("phone", "")
                f_email   = extracted.get("email", "")
                f_id1t    = extracted.get("id1_type", "")
                f_id1n    = extracted.get("id1_number", "")
                f_id2t    = extracted.get("id2_type", "")
                f_id2n    = extracted.get("id2_number", "")
                f_stage   = extracted.get("stage_names", display)

                upsert_signature(
                    shoot_id=shoot.shoot_id,
                    shoot_date=shoot.shoot_date,
                    scene_id=scene_id,
                    studio=studio,
                    talent_role=role,
                    talent_slug=slug,
                    talent_display=display,
                    legal_name=f_legal,
                    business_name=f_biz,
                    tax_classification="individual",
                    llc_class="",
                    other_classification="",
                    exempt_payee_code="",
                    fatca_code="",
                    tin_type="ssn",
                    tin="",
                    dob=f_dob,
                    place_of_birth=f_pob,
                    street_address=f_street,
                    city_state_zip=f_city,
                    phone=f_phone,
                    email=f_email,
                    id1_type=f_id1t,
                    id1_number=f_id1n,
                    id2_type=f_id2t,
                    id2_number=f_id2n,
                    stage_names=f_stage,
                    professional_names="",
                    nicknames_aliases="",
                    previous_legal_names="",
                    signature_image_path=str(sig_path.relative_to(get_settings().base_dir)),
                    signed_ip=request.client.host if request.client else "",
                    signed_user_agent="legacy-import-bulk",
                    signed_by_user=audit_user,
                    pdf_local_path=str(pdf_path),
                    pdf_mega_path=mega_remote if (mega_remote and not mega_err) else "",
                )
                imported += 1

            result.shoots.append(BulkDriveImportShootResult(
                shoot_id=shoot.shoot_id,
                shoot_date=shoot.shoot_date,
                folder_name=name,
                talents_imported=imported,
                skipped_reason="" if imported else "No PDFs matched talent slug prefix",
            ))

    return result


class MegaLegalFile(BaseModel):
    """One file living inside a MEGA ``{SCENE_ID}/Legal/`` folder.

    Source-of-truth for the Compliance "Database" view's *MEGA Legacy* rows —
    paperwork that exists on the bucket but was never imported into
    ``compliance_signatures`` (i.e. no structured row to edit).
    """
    studio: str               # canonical 4-letter code, e.g. "VRH"
    scene_id: str             # canonical scene id, e.g. "VRH0762"
    key: str                  # full bucket key, e.g. "VRH0762/Legal/foo.pdf"
    filename: str             # last path segment
    size: int                 # bytes
    last_modified: str        # ISO8601 UTC


class MegaLegalScanResponse(BaseModel):
    files: list[MegaLegalFile]
    scanned_at: str
    studios_scanned: list[str]
    total: int
    truncated: bool = False


# Module-level cache. Scans are slow (40k+ keys across 4 buckets), so we
# cache the result for 1 hour and let the operator hit "Refresh" on demand.
_MEGA_LEGAL_CACHE: dict[str, tuple[float, MegaLegalScanResponse]] = {}
_MEGA_LEGAL_TTL = 60 * 60   # 1 hour
_MEGA_LEGAL_MAX = 50_000    # safety cap on total files returned


# Extensions we accept as plausible paperwork artifacts. Everything else
# (mp4, mov, ini, ds_store, docx writeups, raw camera files, etc.) is filtered
# out — Legal/ folders historically accumulated everything from BTS videos
# to desktop.ini, and surfacing 10k+ junk rows in the Database view defeated
# the purpose. PDF + the four common ID-photo formats are the real signal.
_PAPERWORK_EXTS = (".pdf", ".jpg", ".jpeg", ".png", ".heic", ".webp")

# Filename patterns that indicate cruft in a Legal/ folder. Skipped
# unconditionally even if the extension matches.
_PAPERWORK_NAME_DENY = (
    "desktop.ini", "thumbs.db", ".ds_store", "icon\r",
)
_PAPERWORK_NAME_DENY_PREFIXES = (
    "copy of ",        # duplicate cruft like "Copy of Copy of foo.jpg"
    "dsc_", "dsc-",    # default camera names
    "img_", "img-",    # generic phone/camera defaults
    "photo ",          # macOS Photos app exports
    "._",              # macOS resource forks
)


def _is_paperwork_filename(filename: str) -> bool:
    """True if this name looks like an actual paperwork artifact rather than
    BTS / desktop / camera junk. The compliance Database view relies on this
    to keep noise out of the merged list."""
    low = filename.lower().strip()
    if not low or low in _PAPERWORK_NAME_DENY:
        return False
    if any(low.startswith(p) for p in _PAPERWORK_NAME_DENY_PREFIXES):
        return False
    if not any(low.endswith(ext) for ext in _PAPERWORK_EXTS):
        return False
    return True


def _scan_mega_legal_folders(
    studios: tuple[str, ...],
    *,
    max_files: int = _MEGA_LEGAL_MAX,
) -> tuple[list[MegaLegalFile], bool]:
    """Walk each studio bucket and yield every key whose path contains
    ``/Legal/`` AND whose filename looks like a real paperwork artifact.

    We can't filter server-side (S3 has no contains-style prefix), so we
    paginate the whole bucket and filter in memory. With ~10k objects per
    studio this completes in seconds; the result is cached for an hour.

    Filename filter: PDFs + the four common ID-photo formats only, with
    obvious cruft (Copy of …, DSC_…, desktop.ini, .DS_Store, etc.) excluded.
    """
    import s4_client

    out: list[MegaLegalFile] = []
    truncated = False
    for studio in studios:
        try:
            for obj in s4_client.list_objects(studio):
                key = obj["key"]
                # Match any nested /Legal/ subfolder (case-insensitive in
                # case some scenes were migrated with "legal/" lowercase).
                if "/legal/" not in key.lower():
                    continue
                # Skip "directory placeholder" keys ending in / with size 0.
                if key.endswith("/"):
                    continue
                filename = key.rsplit("/", 1)[-1]
                # Drop BTS videos, desktop.ini, "Copy of …" cruft, etc.
                if not _is_paperwork_filename(filename):
                    continue
                head, _, _ = key.partition("/")
                try:
                    scene_id = s4_client.normalize_scene_id(head)
                except ValueError:
                    scene_id = head  # non-canonical prefix; surface as-is
                lm = obj.get("last_modified")
                lm_str = ""
                if lm is not None:
                    try:
                        lm_str = lm.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
                    except Exception:
                        lm_str = str(lm)
                out.append(MegaLegalFile(
                    studio=studio.upper(),
                    scene_id=scene_id,
                    key=key,
                    filename=filename,
                    size=int(obj.get("size") or 0),
                    last_modified=lm_str,
                ))
                if len(out) >= max_files:
                    truncated = True
                    break
        except Exception as exc:
            _log.warning("MEGA legal scan failed for %s: %s", studio, exc)
        if truncated:
            break

    out.sort(key=lambda f: (f.scene_id, f.filename))
    return out, truncated


@router.get("/admin/legal-folders", response_model=MegaLegalScanResponse)
async def scan_mega_legal_folders(
    _admin: dict = Depends(require_admin),
    refresh: bool = Query(default=False, description="Force live S3 scan, bypassing the DB index"),
    studio: Optional[str] = Query(default=None, description="Restrict to one studio (FPVR/VRH/VRA/NJOI)"),
):
    """Return the searchable list of MEGA `{SCENE_ID}/Legal/` files.

    Hot path: query the ``compliance_legal_files`` index table that the
    nightly scraper (scrape_mega_legal.py) populates. Sub-millisecond DB
    read instead of a 30+ second paginated bucket walk per page-load.

    Fallback: if the index is empty (first install) or ``refresh=1`` is
    passed, fall back to the live S3 scan + the in-process 1h cache.
    """
    studios: tuple[str, ...] = (studio.lower(),) if studio else ("fpvr", "vrh", "vra", "njoi")

    # Hot path — read from the index. Empty result => fall through to live scan.
    if not refresh:
        with get_db() as conn:
            scan_meta = conn.execute(
                "SELECT last_scan_at FROM compliance_legal_scan_meta WHERE id=1"
            ).fetchone()
            scanned_at = (scan_meta["last_scan_at"] if scan_meta else "") or ""

            if scanned_at:
                where = ""
                params: list = []
                if studio:
                    where = " WHERE studio = ?"
                    params.append(studio.upper())
                rows = conn.execute(
                    f"""SELECT studio, scene_id, key, filename, size, last_modified
                          FROM compliance_legal_files{where}
                         ORDER BY scene_id, filename""",
                    params,
                ).fetchall()

            if scanned_at and rows:
                files = [MegaLegalFile(**dict(r)) for r in rows]
                return MegaLegalScanResponse(
                    files=files,
                    scanned_at=scanned_at,
                    studios_scanned=[s.upper() for s in studios],
                    total=len(files),
                    truncated=False,
                )

    # Cold path / explicit refresh — live S3 scan with 1h in-process cache.
    import time as _time
    cache_key = "|".join(studios)
    now = _time.time()
    cached = _MEGA_LEGAL_CACHE.get(cache_key)
    if cached and not refresh and (now - cached[0]) < _MEGA_LEGAL_TTL:
        return cached[1]

    files, truncated = _scan_mega_legal_folders(studios)
    resp = MegaLegalScanResponse(
        files=files,
        scanned_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        studios_scanned=[s.upper() for s in studios],
        total=len(files),
        truncated=truncated,
    )
    _MEGA_LEGAL_CACHE[cache_key] = (now, resp)
    return resp


class MegaImportRequest(BaseModel):
    """Bulk-import already-signed paperwork sitting in MEGA legal folders.

    Walks every shoot in the date window, lists its ``{SCENE_ID}/Legal/``
    folder, matches each PDF to a talent by filename prefix, downloads the
    bytes, extracts AcroForm fields, and upserts a ``compliance_signatures``
    row. This is the parallel of /admin/bulk-import-from-drive but for
    paperwork that lives natively on MEGA (the migration path going forward).

    Set ``overwrite_existing=True`` to re-import shoots that already have
    rows — useful when a PDF was edited after the initial import.
    """
    date_from: str                        # YYYY-MM-DD inclusive
    date_to: str                          # YYYY-MM-DD inclusive
    studio: Optional[str] = None          # restrict to one canonical studio (FuckPassVR/VRHush/VRAllure/NaughtyJOI)
    overwrite_existing: bool = False
    imported_from_label: str = ""         # recorded in audit trail


class MegaImportShootResult(BaseModel):
    shoot_id: str
    shoot_date: str
    scene_id: str
    studio: str
    talents_imported: int = 0
    talents_skipped: int = 0
    skipped_reason: str = ""


class MegaImportResult(BaseModel):
    shoots_seen: int = 0
    shoots_processed: int = 0
    total_imported: int = 0
    shoots: list[MegaImportShootResult] = []
    errors: list[str] = []


def _download_mega_object_bytes(studio: str, key: str) -> bytes:
    """Pull one object's bytes out of MEGA into memory (no /tmp dance).

    The caller is responsible for sizing — we cap reads at 50MB to protect
    the FastAPI process from a runaway allocation if a Legal/ folder ever
    contains an unexpected mega-file.
    """
    import s4_client
    bucket = s4_client._studio_to_bucket(studio)
    resp = s4_client._client().get_object(Bucket=bucket, Key=key)
    body = resp["Body"]
    MAX_BYTES = 50 * 1024 * 1024
    data = body.read(MAX_BYTES + 1)
    if len(data) > MAX_BYTES:
        raise ValueError(f"Object exceeds {MAX_BYTES} bytes safety cap")
    return data


@router.post("/admin/import-from-mega-legal", response_model=MegaImportResult)
async def import_from_mega_legal(
    request: Request,
    body: MegaImportRequest,
    _admin: dict = Depends(require_admin),
):
    """Bulk-populate ``compliance_signatures`` from MEGA legal folders.

    For each BG/SOLO/JOI shoot in the date window:
      1. Look up the scene's MEGA bucket key prefix (``{SCENE_ID}/Legal/``).
      2. List the PDFs and match each to female/male by filename prefix
         ``{TalentSlug}-`` (the prepare-flow naming convention).
      3. Download bytes via S4, extract AcroForm fields, and upsert one
         signature row per matched talent.

    Skips shoots that already have signature rows unless
    ``overwrite_existing=True``. Errors on individual shoots don't abort
    the batch — they're collected and returned in ``errors``.
    """
    from_d = _parse_shoot_date(body.date_from)
    to_d = _parse_shoot_date(body.date_to)
    if not from_d or not to_d or from_d > to_d:
        raise HTTPException(status_code=400, detail="Invalid date range")

    shoots = _load_shoots_window(from_d, to_d, include_cancelled=False)
    studio_filter = body.studio or ""

    placeholder_png = _import_placeholder_signature_png()
    audit_user = (
        f"legacy_import:mega"
        + (f":{body.imported_from_label}" if body.imported_from_label else "")
    )
    result = MegaImportResult()

    # Pre-check: which shoots already have rows? Cheaper than per-shoot
    # round-tripping the upsert helper.
    existing = list_signed_talents([s.shoot_id for s in shoots])

    for shoot in shoots:
        result.shoots_seen += 1
        bg_scenes = [sc for sc in shoot.scenes if sc.scene_type.upper() in COMPLIANCE_SCENE_TYPES]
        if not bg_scenes:
            continue

        primary = bg_scenes[0]
        scene_id = primary.scene_id or ""
        studio = primary.studio or ""

        if not scene_id or not studio:
            result.shoots.append(MegaImportShootResult(
                shoot_id=shoot.shoot_id, shoot_date=shoot.shoot_date,
                scene_id=scene_id, studio=studio,
                skipped_reason="Missing scene_id or studio",
            ))
            continue
        if studio_filter and studio != studio_filter:
            continue

        prior = existing.get(shoot.shoot_id, [])
        prior_roles = {t.talent_role for t in prior}
        if prior and not body.overwrite_existing:
            result.shoots.append(MegaImportShootResult(
                shoot_id=shoot.shoot_id, shoot_date=shoot.shoot_date,
                scene_id=scene_id, studio=studio,
                talents_skipped=len(prior),
                skipped_reason=f"Already has {len(prior)} row(s); set overwrite to re-import",
            ))
            continue

        # List the Legal/ folder for this scene. Try uppercase first, then
        # the lowercase fallback (the migration left ~23 VRH scenes lowercase).
        import s4_client
        prefix = f"{scene_id}/Legal/"
        files: list[dict] = []
        try:
            files = list(s4_client.list_objects(studio, prefix))
            if not files:
                # lowercase fallback
                files = list(s4_client.list_objects(studio, prefix.lower()))
        except Exception as exc:
            result.errors.append(f"{shoot.shoot_id}: list failed — {exc}")
            continue

        pdfs = [f for f in files if (f["key"] or "").lower().endswith(".pdf")]
        if not pdfs:
            result.shoots.append(MegaImportShootResult(
                shoot_id=shoot.shoot_id, shoot_date=shoot.shoot_date,
                scene_id=scene_id, studio=studio,
                skipped_reason="No PDFs in Legal/ folder",
            ))
            continue

        # Match each PDF to female / male by slug prefix (prepare-flow
        # convention is `{TalentSlug}-{date}.pdf`).
        female_slug = shoot.female_talent.replace(" ", "")
        male_slug = (shoot.male_talent or "").replace(" ", "")
        matches: dict[str, dict] = {}
        for f in pdfs:
            name = (f["key"].rsplit("/", 1)[-1] or "").replace(" ", "")
            low = name.lower()
            if female_slug and low.startswith(female_slug.lower()):
                matches.setdefault("female", f)
            elif male_slug and low.startswith(male_slug.lower()):
                matches.setdefault("male", f)
            elif female_slug and female_slug.lower() in low:
                matches.setdefault("female", f)
            elif male_slug and male_slug.lower() in low:
                matches.setdefault("male", f)

        if not matches:
            result.shoots.append(MegaImportShootResult(
                shoot_id=shoot.shoot_id, shoot_date=shoot.shoot_date,
                scene_id=scene_id, studio=studio,
                skipped_reason=(
                    f"No PDFs matched talent slugs ({female_slug!r}, {male_slug or '—'!r})"
                ),
            ))
            continue

        # If overwriting, skip roles that already have a row of the same
        # talent_slug — we don't want to clobber an actively edited row.
        if body.overwrite_existing:
            for prior_t in prior:
                if prior_t.talent_role in matches and prior_t.talent_slug != (
                    female_slug if prior_t.talent_role == "female" else male_slug
                ):
                    matches.pop(prior_t.talent_role, None)

        shoot_date_obj = _parse_shoot_date(shoot.shoot_date)
        if not shoot_date_obj:
            result.errors.append(f"{shoot.shoot_id}: invalid shoot_date")
            continue
        date_code = shoot_date_obj.strftime("%m%d%y")
        imported_here = 0
        for role, mega_file in matches.items():
            slug = female_slug if role == "female" else male_slug
            display = shoot.female_talent if role == "female" else (shoot.male_talent or "")
            key = mega_file["key"]
            try:
                pdf_bytes = await asyncio.to_thread(_download_mega_object_bytes, studio, key)
            except Exception as exc:
                result.errors.append(f"{shoot.shoot_id}/{role}: download failed — {exc}")
                continue

            # Save locally so render-pdf can serve a stable artifact.
            pdf_name = f"{slug}-{date_code}.pdf"
            pdf_path = _legal_pdf_dir() / shoot.shoot_date / pdf_name
            pdf_path.parent.mkdir(parents=True, exist_ok=True)
            pdf_path.write_bytes(pdf_bytes)

            # Placeholder signature image (we can't recover the original).
            sig_path = _signature_dir() / f"{shoot.shoot_date}-{slug}-{role}.png"
            sig_path.write_bytes(placeholder_png)

            extracted = _extract_pdf_fields_for_import(pdf_bytes)
            mega_remote = (
                f"s3://{s4_client._studio_to_bucket(studio)}/{key}"
            )

            upsert_signature(
                shoot_id=shoot.shoot_id,
                shoot_date=shoot.shoot_date,
                scene_id=scene_id,
                studio=studio,
                talent_role=role,
                talent_slug=slug,
                talent_display=display,
                legal_name=extracted.get("legal_name", display),
                business_name=extracted.get("business_name", ""),
                tax_classification="individual",
                llc_class="",
                other_classification="",
                exempt_payee_code="",
                fatca_code="",
                tin_type="ssn",
                tin="",
                dob=extracted.get("dob", ""),
                place_of_birth=extracted.get("place_of_birth", ""),
                street_address=extracted.get("street_address", ""),
                city_state_zip=extracted.get("city_state_zip", ""),
                phone=extracted.get("phone", ""),
                email=extracted.get("email", ""),
                id1_type=extracted.get("id1_type", ""),
                id1_number=extracted.get("id1_number", ""),
                id2_type=extracted.get("id2_type", ""),
                id2_number=extracted.get("id2_number", ""),
                stage_names=extracted.get("stage_names", display),
                professional_names="",
                nicknames_aliases="",
                previous_legal_names="",
                signature_image_path=str(sig_path.relative_to(get_settings().base_dir)),
                signed_ip=request.client.host if request.client else "",
                signed_user_agent="legacy-import-mega-bulk",
                signed_by_user=audit_user,
                pdf_local_path=str(pdf_path),
                pdf_mega_path=mega_remote,
            )
            imported_here += 1

        result.shoots.append(MegaImportShootResult(
            shoot_id=shoot.shoot_id, shoot_date=shoot.shoot_date,
            scene_id=scene_id, studio=studio,
            talents_imported=imported_here,
            skipped_reason="" if imported_here else "No PDFs matched talent slug",
        ))
        result.shoots_processed += 1
        result.total_imported += imported_here

    # Bust the legal-folders cache so the Database view re-counts MEGA
    # files vs. DB rows on the next refresh.
    _MEGA_LEGAL_CACHE.clear()
    return result


@router.get("/admin/legal-folders/presign")
async def presign_mega_legal_file(
    _admin: dict = Depends(require_admin),
    studio: str = Query(...),
    key: str = Query(...),
):
    """Generate a presigned download URL for one MEGA legal-folder file.

    Used by the Database view when an admin clicks "Open" on a row that has
    only a MEGA artifact (no structured signature row to render). 7-day TTL
    is the SigV4 maximum.
    """
    import s4_client
    if "/legal/" not in key.lower():
        raise HTTPException(status_code=400, detail="Key must be inside a Legal/ folder")
    url = s4_client.presign(studio.lower(), key)
    return {"url": url, "studio": studio.upper(), "key": key}


class SignatureSearchHitOut(BaseModel):
    """Wire format for a single search hit. TIN is masked to last-4 only —
    the full TIN is loaded on demand when an admin opens the edit modal."""
    id: int
    shoot_id: str
    shoot_date: str
    scene_id: str = ""
    studio: str = ""
    talent_role: str
    talent_slug: str
    talent_display: str
    legal_name: str
    business_name: str = ""
    stage_names: str = ""
    nicknames_aliases: str = ""
    previous_legal_names: str = ""
    email: str = ""
    phone: str = ""
    city_state_zip: str = ""
    tin_type: str = ""
    tin_last4: str = ""
    dob: str = ""
    signed_at: str = ""
    pdf_mega_path: str = ""
    contract_version: str = ""


class SignatureSearchResponse(BaseModel):
    """Paginated search response. ``total`` reflects the full match count
    so the UI can render "Showing N of M"."""
    hits: list[SignatureSearchHitOut]
    total: int
    limit: int
    offset: int
    query: str = ""
    date_from: str = ""
    date_to: str = ""
    studio: str = ""
    role: str = ""


@router.get("/admin/search", response_model=SignatureSearchResponse)
async def search_compliance_signatures(
    _admin: dict = Depends(require_admin),
    q: Optional[str] = Query(default=None, description="Free-text query (whitespace-tokenized; AND across tokens)"),
    date_from: Optional[str] = Query(default=None, alias="from"),
    date_to: Optional[str] = Query(default=None, alias="to"),
    studio: Optional[str] = Query(default=None),
    role: Optional[str] = Query(default=None, pattern="^(female|male)$"),
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
):
    """Search every paperwork record we have on file (TKT-paperwork-db).

    Powers the Compliance "Database" view — a flat searchable index of
    every ``compliance_signatures`` row, including legacy Drive imports
    that were back-filled via the bulk importer.

    Free-text matches against talent display/legal/business name, stage
    names, aliases, previous legal names, email, phone, address, scene id,
    shoot id, and studio. TIN is intentionally NOT searchable.

    Admin-only because every row contains personal data.
    """
    hits, total = search_signatures(
        query=q,
        date_from=date_from,
        date_to=date_to,
        studio=studio,
        role=role,
        limit=limit,
        offset=offset,
    )
    return SignatureSearchResponse(
        hits=[SignatureSearchHitOut(**h.__dict__) for h in hits],
        total=total,
        limit=limit,
        offset=offset,
        query=q or "",
        date_from=date_from or "",
        date_to=date_to or "",
        studio=studio or "",
        role=role or "",
    )


@router.get("/admin/w9-summary")
async def w9_summary(
    _admin: dict = Depends(require_admin),
    date_from: Optional[str] = Query(default=None, alias="from"),
    date_to: Optional[str] = Query(default=None, alias="to"),
    studio: Optional[str] = Query(default=None),
):
    """Counts to render in the admin panel before downloading."""
    records = list_w9_records(date_from=date_from, date_to=date_to, studio=studio)
    by_studio: dict[str, int] = {}
    by_role: dict[str, int] = {"female": 0, "male": 0}
    for r in records:
        by_studio[r.studio or "—"] = by_studio.get(r.studio or "—", 0) + 1
        by_role[r.talent_role] = by_role.get(r.talent_role, 0) + 1
    return {
        "total": len(records),
        "by_studio": by_studio,
        "by_role": by_role,
        "date_from": date_from or "",
        "date_to": date_to or "",
        "studio": studio or "",
    }
